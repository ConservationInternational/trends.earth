# -*- coding: utf-8 -*-
"""
/***************************************************************************
 LDMP - A QGIS plugin
 This plugin supports monitoring and reporting of land degradation to the UNCCD 
 and in support of the SDG Land Degradation Neutrality (LDN) target.
                              -------------------
        begin                : 2017-05-23
        git sha              : $Format:%H$
        copyright            : (C) 2017 by Conservation International
        email                : trends.earth@conservation.org
 ***************************************************************************/
"""

import os
import tempfile
import json

from copy import copy

import numpy as np

from osgeo import gdal, osr

import openpyxl
from openpyxl.drawing.image import Image
from openpyxl.styles import Font, Alignment

from qgis.utils import iface
from qgis.core import QgsGeometry
mb = iface.messageBar()

from qgis.PyQt import QtWidgets
from qgis.PyQt.QtCore import QSettings, QDate

from LDMP import log
from LDMP.api import run_script
from LDMP.calculate import DlgCalculateBase, get_script_slug, ClipWorker, \
    json_geom_to_geojson
from LDMP.layers import add_layer, create_local_json_metadata, get_band_infos
from LDMP.worker import AbstractWorker, StartWorker
from LDMP.gui.DlgCalculateRestBiomassData import Ui_DlgCalculateRestBiomassData
from LDMP.gui.DlgCalculateRestBiomassSummaryTable import Ui_DlgCalculateRestBiomassSummaryTable
from LDMP.schemas.schemas import BandInfo, BandInfoSchema
from LDMP.summary import *


class DlgCalculateRestBiomassData(DlgCalculateBase, Ui_DlgCalculateRestBiomassData):
    def __init__(self, parent=None):
        super(DlgCalculateRestBiomassData, self).__init__(parent)

        self.setupUi(self)

        self.first_show = True

    def showEvent(self, event):
        super(DlgCalculateRestBiomassData, self).showEvent(event)

        if self.reset_tab_on_showEvent:
            self.TabBox.setCurrentIndex(0)

    def btn_calculate(self):
        # Note that the super class has several tests in it - if they fail it
        # returns False, which would mean this function should stop execution
        # as well.
        ret = super(DlgCalculateRestBiomassData, self).btn_calculate()
        if not ret:
            return

        self.calculate_on_GEE()

    def get_rest_type(self):
        if self.radioButton_rest_type_terrestrial.isChecked():
            return "terrestrial"
        elif self.radioButton_rest_type_coastal.isChecked():
            return "coastal"
        else:
            # Should never get here
            raise

    def calculate_on_GEE(self):
        self.close()

        crosses_180th, geojsons = self.aoi.bounding_box_gee_geojson()
        payload = {'length_yr': self.spinBox_years.value(),
                   'rest_type': self.get_rest_type(),
                   'geojsons': json.dumps(geojsons),
                   'crs': self.aoi.get_crs_dst_wkt(),
                   'crosses_180th': crosses_180th,
                   'task_name': self.options_tab.task_name.text(),
                   'task_notes': self.options_tab.task_notes.toPlainText()}

        resp = run_script(get_script_slug('restoration-biomass'), payload)

        if resp:
            mb.pushMessage(QtWidgets.QApplication.translate("LDMP", "Submitted"),
                           QtWidgets.QApplication.translate("LDMP", "Restoration biomass change submitted to Google Earth Engine."),
                           level=0, duration=5)
        else:
            mb.pushMessage(QtWidgets.QApplication.translate("LDMP", "Error"),
                           QtWidgets.QApplication.translate("LDMP", "Unable to submit restoration biomass change task to Google Earth Engine."),
                           level=0, duration=5)

class RestBiomassSummaryWorker(AbstractWorker):
    def __init__(self, src_file):
        AbstractWorker.__init__(self)

        self.src_file = src_file

    def work(self):
        self.toggle_show_progress.emit(True)
        self.toggle_show_cancel.emit(True)

        src_ds = gdal.Open(self.src_file)

        band_biomass_initial = src_ds.GetRasterBand(1)
        # First band is initial biomass, and all other bands are for different 
        # types of restoration
        n_types = src_ds.RasterCount - 1

        block_sizes = band_biomass_initial.GetBlockSize()
        xsize = band_biomass_initial.XSize
        ysize = band_biomass_initial.YSize

        x_block_size = block_sizes[0]
        y_block_size = block_sizes[1]

        src_gt = src_ds.GetGeoTransform()

        # Width of cells in longitude
        long_width = src_gt[1]
        # Set initial lat ot the top left corner latitude
        lat = src_gt[3]
        # Width of cells in latitude
        pixel_height = src_gt[5]

        area_site = 0
        biomass_initial = 0
        biomass_change = np.zeros(n_types)

        blocks = 0
        for y in range(0, ysize, y_block_size):
            if y + y_block_size < ysize:
                rows = y_block_size
            else:
                rows = ysize - y
            for x in range(0, xsize, x_block_size):
                if self.killed:
                    log("Processing of {} killed by user after processing {} out of {} blocks.".format(self.prod_out_file, y, ysize))
                    break
                self.progress.emit(100 * (float(y) + (float(x)/xsize)*y_block_size) / ysize)
                if x + x_block_size < xsize:
                    cols = x_block_size
                else:
                    cols = xsize - x

                biomass_initial_array = band_biomass_initial.ReadAsArray(x, y, cols, rows)

                # Caculate cell area for each horizontal line
                cell_areas = np.array([calc_cell_area(lat + pixel_height*n, lat + pixel_height*(n + 1), long_width) for n in range(rows)])
                cell_areas.shape = (cell_areas.size, 1)
                # Make an array of the same size as the input arrays containing 
                # the area of each cell (which is identicalfor all cells in a 
                # given row - cell areas only vary among rows)
                cell_areas_array = np.repeat(cell_areas, cols, axis=1)
                # Convert cell areas to hectares
                cell_areas_array = cell_areas_array * 1e-4

                # The site area includes everything that isn't masked
                site_pixels = biomass_initial_array != -32767
                
                area_site = area_site + np.sum(site_pixels * cell_areas_array)
                biomass_initial = biomass_initial + np.sum(site_pixels * cell_areas_array * biomass_initial_array)

                for n in range(n_types):
                    biomass_rest_array = src_ds.GetRasterBand(n + 2).ReadAsArray(x, y, cols, rows)
                    biomass_change[n] = biomass_change[n] + np.sum((biomass_rest_array) * cell_areas_array * site_pixels)

                blocks += 1
            lat += pixel_height * rows
        self.progress.emit(100)

        if self.killed:
            return None

        return list((biomass_initial, biomass_change, area_site))

class DlgCalculateRestBiomassSummaryTable(DlgCalculateBase, Ui_DlgCalculateRestBiomassSummaryTable):
    def __init__(self, parent=None):
        super(DlgCalculateRestBiomassSummaryTable, self).__init__(parent)

        self.setupUi(self)

        self.add_output_tab(['.json', '.tif', '.xlsx'])

    def showEvent(self, event):
        super(DlgCalculateRestBiomassSummaryTable, self).showEvent(event)

        self.combo_layer_biomass_diff.populate()

    def btn_calculate(self):
        # Note that the super class has several tests in it - if they fail it
        # returns False, which would mean this function should stop execution
        # as well.
        ret = super(DlgCalculateRestBiomassSummaryTable, self).btn_calculate()
        if not ret:
            return

        ######################################################################
        # Check that all needed input layers are selected
        if len(self.combo_layer_biomass_diff.layer_list) == 0:
            QtWidgets.QMessageBox.critical(None, self.tr("Error"),
                                       self.tr("You must add a biomass layer to your map before you can use the summary tool."))
            return
        #######################################################################
        # Check that the layers cover the full extent needed
        if self.aoi.calc_frac_overlap(QgsGeometry.fromRect(self.combo_layer_biomass_diff.get_layer().extent())) < .99:
            QtWidgets.QMessageBox.critical(None, self.tr("Error"),
                                       self.tr("Area of interest is not entirely within the biomass layer."))
            return

        self.close()

        #######################################################################
        # Prep files
        in_file = self.combo_layer_biomass_diff.get_data_file()

        # Remember the first value is an indication of whether dataset is 
        # wrapped across 180th meridian
        wkts = self.aoi.meridian_split('layer', 'wkt', warn=False)[1]
        bbs = self.aoi.get_aligned_output_bounds(in_file)

        output_biomass_diff_tifs = []
        output_biomass_diff_json = self.output_tab.output_basename.text() + '.json'
        for n in range(len(wkts)):
            if len(wkts) > 1:
                output_biomass_diff_tif = os.path.splitext(output_biomass_diff_json)[0] + '_{}.tif'.format(n)
            else:
                output_biomass_diff_tif = os.path.splitext(output_biomass_diff_json)[0] + '.tif'
            output_biomass_diff_tifs.append(output_biomass_diff_tif)

            log(u'Saving clipped biomass file to {}'.format(output_biomass_diff_tif))
            geojson = json_geom_to_geojson(QgsGeometry.fromWkt(wkts[n]).asJson())
            clip_worker = StartWorker(ClipWorker, 'masking layers (part {} of {})'.format(n + 1, len(wkts)), 
                                      in_file, output_biomass_diff_tif, 
                                      geojson, bbs[n])
            if not clip_worker.success:
                QtWidgets.QMessageBox.critical(None, self.tr("Error"),
                                           self.tr("Error masking input layers."))
                return

            ######################################################################
            #  Calculate biomass change summary table
            log('Calculating summary table...')
            rest_summary_worker = StartWorker(RestBiomassSummaryWorker,
                                              'calculating summary table (part {} of {})'.format(n + 1, len(wkts)),
                                              output_biomass_diff_tif)
            if not rest_summary_worker.success:
                QtWidgets.QMessageBox.critical(None, self.tr("Error"),
                                           self.tr("Error calculating biomass change summary table."))
                return
            else:
                if n == 0:
                    biomass_initial, \
                        biomass_change, \
                        area_site = rest_summary_worker.get_return()
                else:
                    this_biomass_initial, \
                        this_biomass_change, \
                        this_area_site = rest_summary_worker.get_return()
                    biomass_initial = biomass_initial + this_biomass_initial
                    biomass_change = biomass_change + this_biomass_change
                    area_site = area_site + this_area_site

        log('area_site: {}'.format(area_site))
        log('biomass_initial: {}'.format(biomass_initial))
        log('biomass_change: {}'.format(biomass_change))

        # Figure out how many years of restoration this data is for, take this 
        # from the second band in the in file
        band_infos = get_band_infos(in_file)
        length_yr = band_infos[1]['metadata']['years']
        # And make a list of the restoration types
        rest_types = [band_info['metadata']['type'] for band_info in band_infos[1:len(band_infos)]]

        make_summary_table(self.output_tab.output_basename.text() + '.xlsx', biomass_initial, 
                           biomass_change, area_site, length_yr, rest_types)

        # Add the biomass_dif layers to the map
        if len(output_biomass_diff_tifs) == 1:
            output_file = output_biomass_diff_tifs[0]
        else:
            output_file = os.path.splitext(output_biomass_diff_json)[0] + '.vrt'
            gdal.BuildVRT(output_file, output_biomass_diff_tifs)
        # Update the band infos to use the masking value (-32767) as the file 
        # no data value, so that stretches are more likely to compute correctly
        for item in band_infos:
            item['no_data_value'] = -32767
        create_local_json_metadata(output_biomass_diff_json, output_file, band_infos,
                                   metadata={'task_name': self.options_tab.task_name.text(),
                                             'task_notes': self.options_tab.task_notes.toPlainText()})
        schema = BandInfoSchema()
        for n in range(1, len(band_infos)):
            add_layer(output_file, n + 1, schema.dump(band_infos[n]))

        return True


def copy_style(a, b):
    b.font = copy(a.font)
    b.fill = copy(a.fill)
    b.border = copy(a.border)
    b.alignment = copy(a.alignment)
    b.number_format = copy(a.number_format)
    b.protection = copy(a.protection)


def make_summary_table(out_file, biomass_initial, biomass_change, area_site, 
                       length_yr, rest_types):

    wb = openpyxl.load_workbook(os.path.join(os.path.dirname(os.path.realpath(__file__)), 'data', 'summary_table_restoration.xlsx'))

    ##########################################################################
    # SDG table
    ws_summary = wb['Restoration Biomass Change']
    ws_summary.cell(6, 2).value = area_site
    ws_summary.cell(7, 2).value = length_yr
    ws_summary.cell(8, 2).value = biomass_initial

    # Insert as many rows as necessary for the number of restoration types, and 
    # copy the styles from the original rows, which will get pushed down
    if len(rest_types) > 1:
        offset = len(rest_types) - 1
        ws_summary.insert_rows(13, offset)
        for n in range(len(rest_types) - 1):
            copy_style(ws_summary.cell(13 + offset, 1), ws_summary.cell(13 + n, 1))
            copy_style(ws_summary.cell(13 + offset, 2), ws_summary.cell(13 + n, 2))
            copy_style(ws_summary.cell(13 + offset, 3), ws_summary.cell(13 + n, 3))

        # Need to remerge cells due to row insertion
        ws_summary.merge_cells(start_row=16 + offset, start_column=1,
                               end_row=16 + offset, end_column=3)
        ws_summary.cell(16 + offset, 1).alignment = Alignment(wrap_text=True)
        ws_summary.row_dimensions[16 + offset].height = 50
        ws_summary.merge_cells(start_row=18 + offset, start_column=1,
                               end_row=18 + offset, end_column=3)
        ws_summary.cell(18 + offset, 1).font = Font(bold=True)
        ws_summary.cell(18 + offset, 1).alignment = Alignment(wrap_text=True)
        ws_summary.row_dimensions[18 + offset].height = 60
        ws_summary.merge_cells(start_row=20 + offset, start_column=1,
                               end_row=20 + offset, end_column=3)
        ws_summary.cell(20 + offset, 1).font = Font(bold=True)
        ws_summary.cell(20 + offset, 1).alignment = Alignment(wrap_text=True)
        ws_summary.row_dimensions[20 + offset].height = 30


    # And write the biomass differences for each restoration type
    for n in range(len(rest_types)):
        ws_summary.cell(13 + n, 1).value = rest_types[n].capitalize()
        ws_summary.cell(13 + n, 2).value = biomass_change[n]
        ws_summary.cell(13 + n, 3).value = biomass_initial + biomass_change[n]

    try:
        ws_summary_logo = Image(os.path.join(os.path.dirname(__file__), 'data', 'trends_earth_logo_bl_300width.png'))
        ws_summary.add_image(ws_summary_logo, 'C1')
    except ImportError:
        # add_image will fail on computers without PIL installed (this will be 
        # an issue on some Macs, likely others). it is only used here to add 
        # our logo, so no big deal.
        pass

    try:
        wb.save(out_file)
        log(u'Summary table saved to {}'.format(out_file))
        QtWidgets.QMessageBox.information(None, QtWidgets.QApplication.translate("LDMP", "Success"),
                                      QtWidgets.QApplication.translate("LDMP", u'Summary table saved to {}'.format(out_file)))

    except IOError:
        log(u'Error saving {}'.format(out_file))
        QtWidgets.QMessageBox.critical(None, QtWidgets.QApplication.translate("LDMP", "Error"),
                                   QtWidgets.QApplication.translate("LDMP", u"Error saving output table - check that {} is accessible and not already open.".format(out_file)))
