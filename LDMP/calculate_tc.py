# -*- coding: utf-8 -*-
"""
/***************************************************************************
 LDMP - A QGIS plugin
 This plugin supports monitoring and reporting of land degradation to the UNCCD 
 and in support of the SDG Land Degradation Neutrality (LDN) target.
                              -------------------
        begin                : 2017-05-23
        git sha              : $Format:%H$
        copyright            : (C) 2017 by Conservation International
        email                : trends.earth@conservation.org
 ***************************************************************************/
"""

from builtins import zip
from builtins import range
import os
import json

import numpy as np

from osgeo import gdal, osr

import openpyxl
from openpyxl.drawing.image import Image

from qgis import processing
from qgis.utils import iface
from qgis.core import (QgsApplication, QgsGeometry, QgsProcessingAlgRunnerTask, 
        QgsTask)

mb = iface.messageBar()

from qgis.PyQt import QtWidgets
from qgis.PyQt.QtCore import QSettings, QDate, QCoreApplication

from LDMP import log, GetTempFilename
from LDMP.api import run_script
from LDMP.calculate import (DlgCalculateBase, get_script_slug, 
        json_geom_to_geojson)
from LDMP.layers import add_layer, create_local_json_metadata
from LDMP.worker import AbstractWorker, StartWorker
from LDMP.gui.DlgCalculateTCData import Ui_DlgCalculateTCData
from LDMP.gui.DlgCalculateTCSummaryTable import Ui_DlgCalculateTCSummaryTable
from LDMP.schemas.schemas import BandInfo, BandInfoSchema
from LDMP.summary import *


class tr_calculate_tc(object):
    def tr(message):
        return QCoreApplication.translate("tr_calculate_tc", message)


# TODO: Still need to code below for local calculation of Total Carbon change
class TCWorker(AbstractWorker):
    def __init__(self, in_vrt, out_f, lc_band_nums, lc_years):
        AbstractWorker.__init__(self)
        self.in_vrt = in_vrt
        self.out_f = out_f
        self.lc_years = lc_years
        self.lc_band_nums = [int(x) for x in lc_band_nums]

    def work(self):
        ds_in = gdal.Open(self.in_vrt)

        soc_band = ds_in.GetRasterBand(1)
        clim_band = ds_in.GetRasterBand(2)

        block_sizes = soc_band.GetBlockSize()
        x_block_size = block_sizes[0]
        y_block_size = block_sizes[1]
        xsize = soc_band.XSize
        ysize = soc_band.YSize

        driver = gdal.GetDriverByName("GTiff")
        # Need a band for SOC degradation, plus bands for annual SOC, and for 
        # annual LC
        ds_out = driver.Create(self.out_f, xsize, ysize,
                1 + len(self.lc_years)*2, gdal.GDT_Int16, 
                               ['COMPRESS=LZW'])
        src_gt = ds_in.GetGeoTransform()
        ds_out.SetGeoTransform(src_gt)
        out_srs = osr.SpatialReference()
        out_srs.ImportFromWkt(ds_in.GetProjectionRef())
        ds_out.SetProjection(out_srs.ExportToWkt())

        blocks = 0
        for y in range(0, ysize, y_block_size):
            if y + y_block_size < ysize:
                rows = y_block_size
            else:
                rows = ysize - y
            for x in range(0, xsize, x_block_size):
                if self.killed:
                    log("Processing of {} killed by user after processing {} out of {} blocks.".format(self.prod_out_file, y, ysize))
                    break
                self.progress.emit(100 * (float(y) + (float(x)/xsize)*y_block_size) / ysize)
                if x + x_block_size < xsize:
                    cols = x_block_size
                else:
                    cols = xsize - x

                # Write initial soc to band 2 of the output file. Read SOC in 
                # as float so the soc change calculations won't accumulate 
                # error due to repeated truncation of ints
                soc = np.array(soc_band.ReadAsArray(x, y, cols, rows)).astype(np.float32)
                ds_out.GetRasterBand(2).WriteArray(soc, x, y)

                blocks += 1

        if self.killed:
            del ds_in
            del ds_out
            os.remove(self.out_f)
            return None
        else:
            return True


class DlgCalculateTCData(DlgCalculateBase, Ui_DlgCalculateTCData):
    def __init__(self, parent=None):
        super(DlgCalculateTCData, self).__init__(parent)

        self.setupUi(self)

        self.first_show = True

    def showEvent(self, event):
        super(DlgCalculateTCData, self).showEvent(event)

        self.use_custom_initial.populate()
        self.use_custom_final.populate()

        self.radioButton_carbon_custom.setEnabled(False)

        if self.reset_tab_on_showEvent:
            self.TabBox.setCurrentIndex(0)

        if self.first_show:
            self.first_show = False
            # Ensure the special value text (set to " ") is displayed by 
            # default
            self.hansen_fc_threshold.setSpecialValueText(' ')
            self.hansen_fc_threshold.setValue(self.hansen_fc_threshold.minimum())

        self.use_hansen.toggled.connect(self.lc_source_changed)
        self.use_custom.toggled.connect(self.lc_source_changed)
        # Ensure that dialogs are enabled/disabled as appropriate
        self.lc_source_changed()

        # Setup bounds for Hansen data
        start_year = QDate(self.datasets['Forest cover']['Hansen']['Start year'], 1, 1)
        end_year = QDate(self.datasets['Forest cover']['Hansen']['End year'], 12, 31)
        self.hansen_bl_year.setMinimumDate(start_year)
        self.hansen_bl_year.setMaximumDate(end_year)
        self.hansen_tg_year.setMinimumDate(start_year)
        self.hansen_tg_year.setMaximumDate(end_year)

    def lc_source_changed(self):
        if self.use_hansen.isChecked():
            self.groupBox_hansen_period.setEnabled(True)
            self.groupBox_hansen_threshold.setEnabled(True)
            self.groupBox_custom_bl.setEnabled(False)
            self.groupBox_custom_tg.setEnabled(False)
        elif self.use_custom.isChecked():
            QtWidgets.QMessageBox.information(None, self.tr("Coming soon!"),
                                       self.tr("Custom forest cover data support is coming soon!"))
            self.use_hansen.setChecked(True)
            # self.groupBox_hansen_period.setEnabled(False)
            # self.groupBox_hansen_threshold.setEnabled(False)
            # self.groupBox_custom_bl.setEnabled(True)
            # self.groupBox_custom_tg.setEnabled(True)


    def get_biomass_dataset(self):
        if self.radioButton_carbon_woods_hole.isChecked():
            return 'woodshole'
        elif self.radioButton_carbon_geocarbon.isChecked():
            return 'geocarbon'
        elif self.radioButton_carbon_custom.isChecked():
            return 'custom'
        else:
            return None

    def get_method(self):
        if self.radioButton_rootshoot_ipcc.isChecked():
            return 'ipcc'
        elif self.radioButton_rootshoot_mokany.isChecked():
            return 'mokany'
        else:
            return None

    def btn_calculate(self):
        # Note that the super class has several tests in it - if they fail it
        # returns False, which would mean this function should stop execution
        # as well.
        ret = super(DlgCalculateTCData, self).btn_calculate()
        if not ret:
            return
        if (self.hansen_fc_threshold.text() == self.hansen_fc_threshold.specialValueText()) and \
                self.use_hansen.isChecked():
            QtWidgets.QMessageBox.critical(None, self.tr("Error"), self.tr(u"Enter a value for percent cover that is considered forest."))
            return

        method = self.get_method()
        if not method:
            QtWidgets.QMessageBox.critical(None, self.tr("Error"), self.tr(u"Choose a method for calculating the root to shoot ratio."))
            return

        biomass_data = self.get_biomass_dataset()
        if not method:
            QtWidgets.QMessageBox.critical(None, self.tr("Error"), self.tr(u"Choose a biomass dataset."))
            return


        if self.use_custom.isChecked():
            self.calculate_locally(method, biomass_data)
        else:
            self.calculate_on_GEE(method, biomass_data)

    def get_save_raster(self):
        raster_file, _ = QtWidgets.QFileDialog.getSaveFileName(self,
                                                        self.tr('Choose a name for the output file'),
                                                        QSettings().value("trends_earth/advanced/base_data_directory", None),
                                                        self.tr('Raster file (*.tif)'))
        if raster_file:
            if os.access(os.path.dirname(raster_file), os.W_OK):
                # QSettings().setValue("LDMP/output_dir", os.path.dirname(raster_file))
                return raster_file
            else:
                QtWidgets.QMessageBox.critical(None, self.tr("Error"),
                                           self.tr(u"Cannot write to {}. Choose a different file.".format(raster_file)))
                return False

    def calculate_locally(self, method, biomass_data):
        if not self.use_custom.isChecked():
            QtWidgets.QMessageBox.critical(None, self.tr("Error"),
                                       self.tr("Due to the options you have chosen, this calculation must occur offline. You MUST select a custom land cover dataset."))
            return


        year_baseline = self.lc_setup_tab.get_initial_year()
        year_target = self.lc_setup_tab.get_final_year()
        if int(year_baseline) >= int(year_target):
            QtWidgets.QMessageBox.information(None, self.tr("Warning"),
                self.tr('The baseline year ({}) is greater than or equal to the target year ({}) - this analysis might generate strange results.'.format(year_baseline, year_target)))

        if self.aoi.calc_frac_overlap(QgsGeometry.fromRect(self.lc_setup_tab.use_custom_initial.get_layer().extent())) < .99:
            QtWidgets.QMessageBox.critical(None, self.tr("Error"),
                                       self.tr("Area of interest is not entirely within the initial land cover layer."))
            return

        if self.aoi.calc_frac_overlap(QgsGeometry.fromRect(self.lc_setup_tab.use_custom_final.get_layer().extent())) < .99:
            QtWidgets.QMessageBox.critical(None, self.tr("Error"),
                                       self.tr("Area of interest is not entirely within the final land cover layer."))
            return

        out_f = self.get_save_raster()
        if not out_f:
            return

        self.close()

        # Select the initial and final bands from initial and final datasets 
        # (in case there is more than one lc band per dataset)
        lc_initial_vrt = self.lc_setup_tab.use_custom_initial.get_vrt()
        lc_final_vrt = self.lc_setup_tab.use_custom_final.get_vrt()
        lc_files = [lc_initial_vrt, lc_final_vrt]
        lc_years = [self.lc_setup_tab.get_initial_year(), self.lc_setup_tab.get_final_year()]
        lc_vrts = []
        for i in range(len(lc_files)):
            f = GetTempFilename('.vrt')
            # Add once since band numbers don't start at zero
            gdal.BuildVRT(f,
                          lc_files[i],
                          bandList=[i + 1],
                          outputBounds=self.aoi.get_aligned_output_bounds_deprecated(lc_initial_vrt),
                          resolution='highest', 
                          resampleAlg=gdal.GRA_NearestNeighbour,
                          separate=True)
            lc_vrts.append(f)

        climate_zones = os.path.join(os.path.dirname(__file__), 'data', 'IPCC_Climate_Zones.tif')
        in_files = [climate_zones]
        in_files.extend(lc_vrts)

        in_vrt = GetTempFilename('.vrt')
        log(u'Saving SOC input files to {}'.format(in_vrt))
        gdal.BuildVRT(in_vrt,
                      in_files,
                      resolution='highest', 
                      resampleAlg=gdal.GRA_NearestNeighbour,
                      outputBounds=self.aoi.get_aligned_output_bounds_deprecated(lc_initial_vrt),
                      separate=True)
        # Lc bands start on band 3 as band 1 is initial soc, and band 2 is 
        # climate zones
        lc_band_nums = np.arange(len(lc_files)) + 3
        # Remove temporary files
        for f in lc_vrts:
            os.remove(f)

        log(u'Saving total carbon to {}'.format(out_f))
        tc_task = StartWorker(TCWorker,
                                 'calculating change in total carbon', 
                                 in_vrt,
                                 out_f,
                                 lc_band_nums,
                                 lc_years)
        os.remove(in_vrt)

        if not tc_task.success:
            QtWidgets.QMessageBox.critical(None, self.tr("Error"),
                                       self.tr("Error calculating change in toal carbon."))
            return

        band_infos = [BandInfo("Total carbon (change)", add_to_map=True, metadata={'year_start': lc_years[0], 'year_end': lc_years[-1]})]
        for year in lc_years:
            if (year == lc_years[0]) or (year == lc_years[-1]):
                # Add first and last years to map
                add_to_map = True
            else:
                add_to_map = False
            band_infos.append(BandInfo("Total carbon", add_to_map=add_to_map, metadata={'year': year}))
        for year in lc_years:
            band_infos.append(BandInfo("Land cover (7 class)", metadata={'year': year}))

        out_json = os.path.splitext(out_f)[0] + '.json'
        create_local_json_metadata(out_json, out_f, band_infos,
                                    metadata={'task_name': self.options_tab.task_name.text(),
                                             'task_notes': self.options_tab.task_notes.toPlainText(),
                                             'source': self.get_subclass_name(),   # linked to calculate.local_script
                                             'id': self.output_tab.process_id,
                                             'start_date': self.output_tab.process_datetime_str})
        schema = BandInfoSchema()
        for band_number in range(len(band_infos)):
            b = schema.dump(band_infos[band_number])
            if b['add_to_map']:
                # The +1 is because band numbers start at 1, not zero
                add_layer(out_f, band_number + 1, b)

    def calculate_on_GEE(self, method, biomass_data):
        self.close()

        crosses_180th, geojsons = self.gee_bounding_box
        payload = {'year_start': self.hansen_bl_year.date().year(),
                   'year_end': self.hansen_tg_year.date().year(),
                   'fc_threshold': int(self.hansen_fc_threshold.text().replace('%', '')),
                   'method': method,
                   'biomass_data': biomass_data,
                   'geojsons': json.dumps(geojsons),
                   'crs': self.aoi.get_crs_dst_wkt(),
                   'crosses_180th': crosses_180th,
                   'task_name': self.options_tab.task_name.text(),
                   'task_notes': self.options_tab.task_notes.toPlainText()}

        resp = run_script(get_script_slug('total-carbon'), payload)

        if resp:
            mb.pushMessage(self.tr("Submitted"),
                           self.tr("Total carbon submitted to Google Earth Engine."),
                           level=0, duration=5)
        else:
            mb.pushMessage(self.tr("Error"),
                           self.tr("Unable to submit total carbon task to Google Earth Engine."),
                           level=0, duration=5)


class TCSummaryWorker(AbstractWorker):
    def __init__(self, src_file, year_start, year_end):
        AbstractWorker.__init__(self)

        self.src_file = src_file
        self.year_start = year_start
        self.year_end = year_end

    def work(self):
        self.toggle_show_progress.emit(True)
        self.toggle_show_cancel.emit(True)

        src_ds = gdal.Open(self.src_file)

        band_f_loss = src_ds.GetRasterBand(1)
        band_tc = src_ds.GetRasterBand(2)

        block_sizes = band_f_loss.GetBlockSize()
        xsize = band_f_loss.XSize
        ysize = band_f_loss.YSize
        n_out_bands = 1

        x_block_size = block_sizes[0]
        y_block_size = block_sizes[1]

        src_gt = src_ds.GetGeoTransform()

        # Width of cells in longitude
        long_width = src_gt[1]
        # Set initial lat ot the top left corner latitude
        lat = src_gt[3]
        # Width of cells in latitude
        pixel_height = src_gt[5]

        area_missing = 0
        area_non_forest = 0
        area_water = 0
        area_site = 0
        initial_forest_area = 0
        initial_carbon_total = 0
        forest_loss = np.zeros(self.year_end - self.year_start)
        carbon_loss = np.zeros(self.year_end - self.year_start)

        blocks = 0
        for y in range(0, ysize, y_block_size):
            if y + y_block_size < ysize:
                rows = y_block_size
            else:
                rows = ysize - y
            for x in range(0, xsize, x_block_size):
                if self.killed:
                    log("Processing of {} killed by user after processing {} out of {} blocks.".format(self.prod_out_file, y, ysize))
                    break
                self.progress.emit(100 * (float(y) + (float(x)/xsize)*y_block_size) / ysize)
                if x + x_block_size < xsize:
                    cols = x_block_size
                else:
                    cols = xsize - x

                f_loss_array = band_f_loss.ReadAsArray(x, y, cols, rows)
                tc_array = band_tc.ReadAsArray(x, y, cols, rows)

                # Caculate cell area for each horizontal line
                cell_areas = np.array([calc_cell_area(lat + pixel_height*n, lat + pixel_height*(n + 1), long_width) for n in range(rows)])
                cell_areas.shape = (cell_areas.size, 1)
                # Make an array of the same size as the input arrays containing 
                # the area of each cell (which is identicalfor all cells ina 
                # given row - cell areas only vary among rows)
                cell_areas_array = np.repeat(cell_areas, cols, axis=1)

                initial_forest_pixels = (f_loss_array == 0) | (f_loss_array > (self.year_start - 2000))
                # The site area includes everything that isn't masked
                area_missing = area_missing + np.sum(((f_loss_array == -32768) | (tc_array == -32768)) * cell_areas_array)
                area_water = area_water + np.sum((f_loss_array == -2) * cell_areas_array)
                area_non_forest = area_non_forest + np.sum((f_loss_array == -1) * cell_areas_array)
                area_site = area_site + np.sum((f_loss_array != -32767) * cell_areas_array)
                initial_forest_area = initial_forest_area + np.sum(initial_forest_pixels * cell_areas_array)
                initial_carbon_total = initial_carbon_total +  np.sum(initial_forest_pixels * tc_array * (tc_array >= 0) * cell_areas_array)

                for n in range(self.year_end - self.year_start):
                    # Note the codes are year - 2000
                    forest_loss[n] = forest_loss[n] + np.sum((f_loss_array == self.year_start - 2000 + n + 1) * cell_areas_array)
                    # Check units here - is tc_array in per m or per ha?
                    carbon_loss[n] = carbon_loss[n] + np.sum((f_loss_array == self.year_start - 2000 + n + 1) * tc_array * (tc_array >= 0) * cell_areas_array)

                blocks += 1
            lat += pixel_height * rows
        self.progress.emit(100)

        if self.killed:
            return None
        else:
            # Convert all area tables from meters into hectares
            forest_loss = forest_loss * 1e-4
            # Note that carbon is scaled by 10
            carbon_loss = carbon_loss * 1e-4 / 10
            area_missing = area_missing * 1e-4
            area_water = area_water * 1e-4
            area_non_forest = area_non_forest * 1e-4
            area_site = area_site * 1e-4
            initial_forest_area = initial_forest_area * 1e-4
            # Note that carbon is scaled by 10
            initial_carbon_total = initial_carbon_total * 1e-4 / 10

        return list((forest_loss, carbon_loss, area_missing, area_water, 
                     area_non_forest, area_site, initial_forest_area, 
                     initial_carbon_total))


def write_excel_summary(forest_loss, carbon_loss, area_missing, area_water, 
                       area_non_forest, area_site, area_forest, 
                       initial_carbon_total, year_start, year_end, out_file):
                          
    wb = openpyxl.load_workbook(os.path.join(os.path.dirname(os.path.realpath(__file__)), 'data', 'summary_table_tc.xlsx'))

    ##########################################################################
    # SDG table
    ws_summary = wb['Total Carbon Summary Table']
    ws_summary.cell(8, 3).value = area_forest
    ws_summary.cell(9, 3).value = area_non_forest
    ws_summary.cell(10, 3).value = area_water
    ws_summary.cell(11, 3).value = area_missing
    ws_summary.cell(15, 4).value = year_start
    ws_summary.cell(16, 4).value = year_end
    #ws_summary.cell(10, 3).value = area_site

    ws_summary.cell(8, 5).value = initial_carbon_total
    write_col_to_sheet(ws_summary, np.arange(year_start + 1, year_end + 1), 1, 24) # Years
    write_col_to_sheet(ws_summary, forest_loss, 2, 24) # Years
    write_col_to_sheet(ws_summary, carbon_loss, 4, 24) # Years

    try:
        ws_summary_logo = Image(os.path.join(os.path.dirname(__file__), 'data', 'trends_earth_logo_bl_300width.png'))
        ws_summary.add_image(ws_summary_logo, 'E1')
    except ImportError:
        # add_image will fail on computers without PIL installed (this will be 
        # an issue on some Macs, likely others). it is only used here to add 
        # our logo, so no big deal.
        pass

    try:
        wb.save(out_file)

    except IOError:
        log(u'Error saving {}'.format(out_file))
        return False

    log(u'Summary table saved to {}'.format(out_file))
    return True

class SummaryTask(QgsTask):
    def __init__(self, aoi, year_start, year_end, f_loss_vrt, tc_vrt, 
            output_file):
        super().__init__('Total carbon summary table calculation', QgsTask.CanCancel)
        
        self.aoi = aoi
        self.year_start = year_start
        self.year_end = year_end
        self.f_loss_vrt = f_loss_vrt
        self.tc_vrt = tc_vrt
        self.output_file = output_file

    def run(self):
        # Remember the first value is an indication of whether dataset is 
        # wrapped across 180th meridian
        wkts = self.aoi.meridian_split('layer', 'wkt', warn=False)[1]
        bbs = self.aoi.get_aligned_output_bounds(self.f_loss_vrt)

        for n in range(len(wkts)):
            if self.isCanceled():
                return False
            ######################################################################
            #  Clip layers
            
            # Combines SDG 15.3.1 input raster into a VRT and crop to the AOI
            indic_vrt = GetTempFilename('.vrt')
            log(u'Saving indicator VRT to: {}'.format(indic_vrt))
            # The plus one is because band numbers start at 1, not zero
            gdal.BuildVRT(indic_vrt,
                          [self.f_loss_vrt, self.tc_vrt],
                          outputBounds=bbs[n],
                          resolution='highest',
                          resampleAlg=gdal.GRA_NearestNeighbour,
                          separate=True)

            clipped_vrt = GetTempFilename('.tif')
            log(u'Saving forest loss/carbon clipped file to {}'.format(clipped_vrt))
            #clip_task = QgsProcessingAlgRunnerTask(
            clip_task = processing.run(
                    'trendsearth:raster_clip',
                    {
                        'INPUT': indic_vrt,
                        'GEOJSON': json.dumps(json_geom_to_geojson(QgsGeometry.fromWkt(wkts[n]).asJson())),
                        'OUTPUT_BOUNDS': str(bbs[n]).strip('[]'),
                        'OUTPUT': clipped_vrt
                    })
            #clip_task.run()
                    # 'masking layers (part {} of {})'.format(n + 1, len(wkts))
                    #
            if self.isCanceled():
                return False
            if not clip_task['SUCCESS']:
                self.exception = Exception('Clipping failed')
                return False

            ######################################################################
            #  Calculate carbon change table
            log('Calculating summary table...')
            summary_task = processing.run(
                    'trendsearth:carbon_summary',
                    {
                        'INPUT': clipped_vrt,
                        'YEAR_START': self.year_start,
                        'YEAR_END': self.year_end
                    })
                    # 'calculating summary table (part {} of {})'.format(n + 1, 
                    # len(wkts))
            if self.isCanceled():
                return
            if not clip_task['SUCCESS']:
                self.exception = Exception('Summarizing carbon change failed')
                return False

            os.remove(indic_vrt)
            os.remove(clipped_vrt)
            os.remove(self.tc_vrt)
            os.remove(self.f_loss_vrt)

            if n == 0:
                 forest_loss = np_array_from_str(summary_task['FOREST_LOSS'])
                 carbon_loss = np_array_from_str(summary_task['CARBON_LOSS'])
                 initial_carbon_total = summary_task['CARBON_INITIAL']
                 area_forest = summary_task['AREA_FOREST']
                 area_non_forest = summary_task['AREA_NON_FOREST']
                 area_water = summary_task['AREA_WATER']
                 area_missing = summary_task['AREA_MISSING']
                 area_site = summary_task['AREA_SITE']
            else:

                 forest_loss = forest_loss + np_array_from_str(summary_task['FOREST_LOSS'])
                 carbon_loss = carbon_loss + np_array_from_str(summary_task['CARBON_LOSS'])
                 area_forest = area_forest + summary_task['AREA_FOREST']
                 area_non_forest = area_non_forest + summary_task['AREA_NON_FOREST']
                 area_water = area_water +  summary_task['AREA_WATER']
                 area_missing = area_missing + summary_task['AREA_MISSING']
                 area_site = area_site + summary_task['AREA_SITE']
                 initial_carbon_total = initial_carbon_total + summary_task['CARBON_INITIAL']

        log('area_missing: {}'.format(area_missing))
        log('area_water: {}'.format(area_water))
        log('area_non_forest: {}'.format(area_non_forest))
        log('area_site: {}'.format(area_site))
        log('area_forest: {}'.format(area_forest))
        log('initial_carbon_total: {}'.format(initial_carbon_total))
        log('forest loss: {}'.format(forest_loss))
        log('carbon loss: {}'.format(carbon_loss))

        write_excel_summary(forest_loss, carbon_loss, area_missing, area_water, 
                           area_non_forest, area_site, area_forest, 
                           initial_carbon_total, self.year_start, 
                           self.year_end, self.output_file)
        return True

    def finished(self, result):
        if self.isCanceled():
            return
        elif result:
            QtWidgets.QMessageBox.information(None, tr_calculate_tc.tr("Success"),
                                          tr_calculate_tc.tr(u'Summary table saved to {}'.format(self.outout_file)))
        else:
            QtWidgets.QMessageBox.critical(None, tr_calculate_tc.tr("Error"),
                                       tr_calculate_tc.tr(u"Error saving output table - check that {} is accessible and not already open.".format(self.output_file)))

class DlgCalculateTCSummaryTable(DlgCalculateBase, Ui_DlgCalculateTCSummaryTable):
    def __init__(self, parent=None):
        super(DlgCalculateTCSummaryTable, self).__init__(parent)

        self.setupUi(self)

        self.add_output_tab(['.xlsx'])

    def showEvent(self, event):
        super(DlgCalculateTCSummaryTable, self).showEvent(event)

        self.combo_layer_f_loss.populate()
        self.combo_layer_tc.populate()

    def btn_calculate(self):
        # Note that the super class has several tests in it - if they fail it
        # returns False, which would mean this function should stop execution
        # as well.
        ret = super(DlgCalculateTCSummaryTable, self).btn_calculate()
        if not ret:
            return

        ######################################################################
        # Check that all needed input layers are selected
        if len(self.combo_layer_f_loss.layer_list) == 0:
            QtWidgets.QMessageBox.critical(None, self.tr("Error"),
                                       self.tr("You must add a forest loss layer to your map before you can use the carbon change summary tool."))
            return
        if len(self.combo_layer_tc.layer_list) == 0:
            QtWidgets.QMessageBox.critical(None, self.tr("Error"),
                                       self.tr("You must add a total carbon layer to your map before you can use the carbon change summary tool."))
            return
        #######################################################################
        # Check that the layers cover the full extent needed
        if self.aoi.calc_frac_overlap(QgsGeometry.fromRect(self.combo_layer_f_loss.get_layer().extent())) < .99:
            QtWidgets.QMessageBox.critical(None, self.tr("Error"),
                                       self.tr("Area of interest is not entirely within the forest loss layer."))
            return
        if self.aoi.calc_frac_overlap(QgsGeometry.fromRect(self.combo_layer_tc.get_layer().extent())) < .99:
            QtWidgets.QMessageBox.critical(None, self.tr("Error"),
                                       self.tr("Area of interest is not entirely within the total carbon layer."))
            return

        #######################################################################
        # Check that all of the productivity layers have the same resolution 
        # and CRS
        def res(layer):
            return (round(layer.rasterUnitsPerPixelX(), 10), round(layer.rasterUnitsPerPixelY(), 10))

        if res(self.combo_layer_f_loss.get_layer()) != res(self.combo_layer_tc.get_layer()):
            QtWidgets.QMessageBox.critical(None, self.tr("Error"),
                                       self.tr("Resolutions of forest loss and total carbon layers do not match."))
            return

        self.close()

        # Load all datasets to VRTs (to select only the needed bands)
        f_loss_vrt = self.combo_layer_f_loss.get_vrt()
        tc_vrt = self.combo_layer_tc.get_vrt()

        # Figure out start and end dates
        year_start = self.combo_layer_f_loss.get_band_info()['metadata']['year_start']
        year_end = self.combo_layer_f_loss.get_band_info()['metadata']['year_end']

        summary_task = SummaryTask(self.aoi, year_start, year_end, 
                f_loss_vrt, tc_vrt, self.output_tab.output_basename.text() + '.xlsx')
        log("Adding task to task manager")
        QgsApplication.taskManager().addTask(summary_task)
        if summary_task.status() not in [QgsTask.Complete, QgsTask.Terminated]:
            QCoreApplication.processEvents()
        # while QgsApplication.taskManager().countActiveTasks() > 0:
        #         QCoreApplication.processEvents()

        return True
