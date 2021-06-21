import datetime as dt
from copy import copy
from pathlib import Path

import numpy as np
import openpyxl
import qgis.core
from openpyxl.styles import (
    Alignment,
    Font,
)
from osgeo import gdal

import LDMP.logger
from .. import (
    areaofinterest,
    calculate,
    calculate_rest_biomass,
    summary,
    utils,
    worker,
)
from ..jobs import models


def compute_biomass_restoration(
        biomass_job: models.Job, area_of_interest: areaofinterest.AOI) -> models.Job:
    #######################################################################
    # Prep files
    in_file = biomass_job.params.params["in_file_path"]

    # Remember the first value is an indication of whether dataset is
    # wrapped across 180th meridian
    _, wkts = area_of_interest.meridian_split('layer', 'wkt', warn=False)
    bbs = area_of_interest.get_aligned_output_bounds(in_file)
    job_output_path, _ = utils.get_local_job_output_paths(biomass_job)
    output_biomass_diff_tifs = []
    for n in range(len(wkts)):
        if len(wkts) > 1:
            output_biomass_diff_tif = (
                    job_output_path.parent / f"{job_output_path.stem}_{n}.tif")
        else:
            output_biomass_diff_tif = (
                    job_output_path.parent / f"{job_output_path.stem}.tif")
        output_biomass_diff_tifs.append(output_biomass_diff_tif)

        LDMP.logger.log(u'Saving clipped biomass file to {}'.format(output_biomass_diff_tif))
        geojson = calculate.json_geom_to_geojson(
            qgis.core.QgsGeometry.fromWkt(wkts[n]).asJson())
        clip_worker = worker.StartWorker(
            calculate.ClipWorker,
            f'masking layers (part {n+1} of {len(wkts)})',
            in_file,
            str(output_biomass_diff_tif),
            geojson,
            bbs[n]
        )
        if clip_worker.success:
            ######################################################################
            #  Calculate biomass change summary table
            LDMP.logger.log('Calculating summary table...')
            rest_summary_worker = worker.StartWorker(
                calculate_rest_biomass.RestBiomassSummaryWorker,
                f'calculating summary table (part {n+1} of {len(wkts)})',
                str(output_biomass_diff_tif)
            )
            if rest_summary_worker.success:
                worker_result = rest_summary_worker.get_return()
                if n == 0:
                    biomass_initial, biomass_change, area_site = worker_result
                else:
                    this_initial, this_change, this_area_site = worker_result
                    biomass_initial = biomass_initial + this_initial
                    biomass_change = biomass_change + this_change
                    area_site = area_site + this_area_site
            else:
                raise RuntimeError("Error calculating biomass change summary table.")


        else:
            raise RuntimeError("Error masking input layers.")

    LDMP.logger.log(f'area_site: {area_site}')
    LDMP.logger.log(f'biomass_initial: {biomass_initial}')
    LDMP.logger.log(f'biomass_change: {biomass_change}')

    summary_table_output_path = job_output_path.parent / f"{job_output_path.stem}.xlsx"
    _save_summary_table(
        str(summary_table_output_path),
        biomass_initial,
        biomass_change,
        area_site,
        biomass_job.params.params["restoration_years"],
        biomass_job.params.params["restoration_types"]
    )
    biomass_job.end_date = dt.datetime.now(dt.timezone.utc)
    biomass_job.progress = 100

    # Add the biomass_dif layers to the map
    if len(output_biomass_diff_tifs) == 1:
        output_path = Path(output_biomass_diff_tifs[0])
    else:
        output_path = job_output_path.parent / f"{job_output_path.stem}.vrt"
        gdal.BuildVRT(str(output_path), output_biomass_diff_tifs)
    biomass_job.results.local_paths = [
        output_path,
        summary_table_output_path,
    ]

    # Update the band infos to use the masking value (-32767) as the file
    # no data value, so that stretches are more likely to compute correctly
    output_bands = []
    for raw_band_info in biomass_job.params.params["in_file_band_infos"]:
        band = models.JobBand.deserialize(raw_band_info)
        band.no_data_value = -32767
        output_bands.append(band)
    biomass_job.results.bands = output_bands
    return biomass_job


def _save_summary_table(
        out_file,
        biomass_initial,
        biomass_change,
        area_site,
        length_yr,
        rest_types
):
    template_summary_table_path = Path(
        __file__).parents[1] / "data/summary_table_restoration.xlsx"
    workbook = openpyxl.load_workbook(str(template_summary_table_path))

    ws_summary_sheet = workbook['Restoration Biomass Change']
    ws_summary_sheet.cell(6, 2).value = area_site
    ws_summary_sheet.cell(7, 2).value = length_yr
    ws_summary_sheet.cell(8, 2).value = biomass_initial

    # Insert as many rows as necessary for the number of restoration types, and
    # copy the styles from the original rows, which will get pushed down
    if len(rest_types) > 1:
        offset = len(rest_types) - 1
        ws_summary_sheet.insert_rows(13, offset)
        for n in range(len(rest_types) - 1):
            _copy_style(
                ws_summary_sheet.cell(13 + offset, 1), ws_summary_sheet.cell(13 + n, 1))
            _copy_style(
                ws_summary_sheet.cell(13 + offset, 2), ws_summary_sheet.cell(13 + n, 2))
            _copy_style(
                ws_summary_sheet.cell(13 + offset, 3), ws_summary_sheet.cell(13 + n, 3))

        # Need to remerge cells due to row insertion
        ws_summary_sheet.merge_cells(
            start_row=16 + offset,
            start_column=1,
            end_row=16 + offset,
            end_column=3
        )
        ws_summary_sheet.cell(16 + offset, 1).alignment = Alignment(wrap_text=True)
        ws_summary_sheet.row_dimensions[16 + offset].height = 50
        ws_summary_sheet.merge_cells(
            start_row=18 + offset,
            start_column=1,
            end_row=18 + offset,
            end_column=3
        )
        ws_summary_sheet.cell(18 + offset, 1).font = Font(bold=True)
        ws_summary_sheet.cell(18 + offset, 1).alignment = Alignment(wrap_text=True)
        ws_summary_sheet.row_dimensions[18 + offset].height = 60
        ws_summary_sheet.merge_cells(
            start_row=20 + offset,
            start_column=1,
            end_row=20 + offset,
            end_column=3
        )
        ws_summary_sheet.cell(20 + offset, 1).font = Font(bold=True)
        ws_summary_sheet.cell(20 + offset, 1).alignment = Alignment(wrap_text=True)
        ws_summary_sheet.row_dimensions[20 + offset].height = 30


    # And write the biomass differences for each restoration type
    for n in range(len(rest_types)):
        ws_summary_sheet.cell(13 + n, 1).value = rest_types[n].capitalize()
        ws_summary_sheet.cell(13 + n, 2).value = biomass_change[n]
        ws_summary_sheet.cell(13 + n, 3).value = biomass_initial + biomass_change[n]

    utils.maybe_add_image_to_sheet(
        "trends_earth_logo_bl_300width.png", ws_summary_sheet)

    try:
        workbook.save(out_file)
        LDMP.logger.log(u'Summary table saved to {}'.format(out_file))
    except IOError as exc:
        raise RuntimeError(
            f"Error saving output table - check that {out_file} is accessible and "
            f"not already open. - {str(exc)}"
        )


def _copy_style(a, b):
    b.font = copy(a.font)
    b.fill = copy(a.fill)
    b.border = copy(a.border)
    b.alignment = copy(a.alignment)
    b.number_format = copy(a.number_format)
    b.protection = copy(a.protection)


class RestBiomassSummaryWorker(worker.AbstractWorker):
    def __init__(self, src_file):
        worker.AbstractWorker.__init__(self)

        self.src_file = src_file

    def work(self):
        self.toggle_show_progress.emit(True)
        self.toggle_show_cancel.emit(True)

        src_ds = gdal.Open(self.src_file)

        band_biomass_initial = src_ds.GetRasterBand(1)
        # First band is initial biomass, and all other bands are for different
        # types of restoration
        n_types = src_ds.RasterCount - 1

        block_sizes = band_biomass_initial.GetBlockSize()
        xsize = band_biomass_initial.XSize
        ysize = band_biomass_initial.YSize

        x_block_size = block_sizes[0]
        y_block_size = block_sizes[1]

        src_gt = src_ds.GetGeoTransform()

        # Width of cells in longitude
        long_width = src_gt[1]
        # Set initial lat ot the top left corner latitude
        lat = src_gt[3]
        # Width of cells in latitude
        pixel_height = src_gt[5]

        area_site = 0
        biomass_initial = 0
        biomass_change = np.zeros(n_types)

        blocks = 0
        for y in range(0, ysize, y_block_size):
            if y + y_block_size < ysize:
                rows = y_block_size
            else:
                rows = ysize - y
            for x in range(0, xsize, x_block_size):
                if self.killed:
                    LDMP.logger.log("Processing of {} killed by user after processing {} out of {} blocks.".format(
                        self.prod_out_file, y, ysize))
                    break
                self.progress.emit(100 * (float(y) + (float(x) / xsize) * y_block_size) / ysize)
                if x + x_block_size < xsize:
                    cols = x_block_size
                else:
                    cols = xsize - x

                biomass_initial_array = band_biomass_initial.ReadAsArray(x, y, cols, rows)

                # Caculate cell area for each horizontal line
                cell_areas = np.array(
                    [summary.calc_cell_area(lat + pixel_height * n, lat + pixel_height * (n + 1), long_width) for n in
                     range(rows)])
                cell_areas.shape = (cell_areas.size, 1)
                # Make an array of the same size as the input arrays containing
                # the area of each cell (which is identicalfor all cells in a
                # given row - cell areas only vary among rows)
                cell_areas_array = np.repeat(cell_areas, cols, axis=1)
                # Convert cell areas to hectares
                cell_areas_array = cell_areas_array * 1e-4

                # The site area includes everything that isn't masked
                site_pixels = biomass_initial_array != -32767

                area_site = area_site + np.sum(site_pixels * cell_areas_array)
                biomass_initial = biomass_initial + np.sum(site_pixels * cell_areas_array * biomass_initial_array)

                for n in range(n_types):
                    biomass_rest_array = src_ds.GetRasterBand(n + 2).ReadAsArray(x, y, cols, rows)
                    biomass_change[n] = biomass_change[n] + np.sum(
                        (biomass_rest_array) * cell_areas_array * site_pixels)

                blocks += 1
            lat += pixel_height * rows
        self.progress.emit(100)

        if self.killed:
            return None

        return list((biomass_initial, biomass_change, area_site))
