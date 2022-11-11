import datetime as dt
import json
import os
from pathlib import Path

import numpy as np
import openpyxl
import processing
import qgis.core
from openpyxl.drawing.image import Image
from osgeo import gdal
from qgis.PyQt import QtCore
from qgis.PyQt import QtWidgets
from te_schemas.results import FileResults
from te_schemas.results import URI

from .. import areaofinterest
from .. import calculate
from .. import GetTempFilename
from .. import logger
from .. import summary
from .. import utils
from ..jobs.models import Job


class tr_calculate_tc:
    def tr(message):
        return QtCore.QCoreApplication.translate("tr_calculate_tc", message)


def compute_total_carbon_summary_table(
    tc_job: Job,
    area_of_interest: areaofinterest.AOI,
    job_output_path: Path,
    dataset_output_path: Path,
) -> Job:
    # Load all datasets to VRTs (to select only the needed bands)
    f_loss_vrt = utils.save_vrt(
        tc_job.params["f_loss_path"], tc_job.params["f_loss_band_index"]
    )
    tc_vrt = utils.save_vrt(tc_job.params["tc_path"], tc_job.params["tc_band_index"])

    summary_table_output_path = job_output_path.parent / f"{job_output_path.stem}.xlsx"
    summary_task = SummaryTask(
        area_of_interest,
        tc_job.params["year_initial"],
        tc_job.params["year_final"],
        f_loss_vrt,
        tc_vrt,
        str(summary_table_output_path),
    )
    logger.log("Adding task to task manager...")
    qgis.core.QgsApplication.taskManager().addTask(summary_task)
    terminal_statuses = [qgis.core.QgsTask.Complete, qgis.core.QgsTask.Terminated]
    if summary_task.status() not in terminal_statuses:
        QtCore.QCoreApplication.processEvents()

    tc_job.results = FileResults(
        name="total_carbon_summary",
        uri=URI(uri=summary_table_output_path),
    )
    tc_job.end_date = dt.datetime.now(dt.timezone.utc)
    tc_job.progress = 100

    return tc_job


class SummaryTask(qgis.core.QgsTask):
    def __init__(self, aoi, year_initial, year_final, f_loss_vrt, tc_vrt, output_file):
        super().__init__(
            "Total carbon summary table calculation", qgis.core.QgsTask.CanCancel
        )

        self.aoi = aoi
        self.year_initial = year_initial
        self.year_final = year_final
        self.f_loss_vrt = f_loss_vrt
        self.tc_vrt = tc_vrt
        self.output_file = output_file

    def run(self):
        # Remember the first value is an indication of whether dataset is
        # wrapped across 180th meridian
        wkts = self.aoi.meridian_split("layer", "wkt", warn=False)[1]
        bbs = self.aoi.get_aligned_output_bounds(self.f_loss_vrt)

        for n in range(len(wkts)):
            if self.isCanceled():
                return False
            ######################################################################
            #  Clip layers

            # Combines SDG 15.3.1 input raster into a VRT and crop to the AOI
            indic_vrt = GetTempFilename(".vrt")
            logger.log("Saving indicator VRT to: {}".format(indic_vrt))
            # The plus one is because band numbers start at 1, not zero
            gdal.BuildVRT(
                indic_vrt,
                [self.f_loss_vrt, self.tc_vrt],
                outputBounds=bbs[n],
                resolution="highest",
                resampleAlg=gdal.GRA_NearestNeighbour,
                separate=True,
            )

            clipped_vrt = GetTempFilename(".tif")
            logger.log(
                "Saving forest loss/carbon clipped file to {}".format(clipped_vrt)
            )
            # clip_task = qgis.core.QgsProcessingAlgRunnerTask(
            clip_task = processing.run(
                "trendsearth:raster_clip",
                {
                    "INPUT": indic_vrt,
                    "GEOJSON": json.dumps(
                        calculate.json_geom_to_geojson(
                            qgis.core.QgsGeometry.fromWkt(wkts[n]).asJson()
                        )
                    ),
                    "OUTPUT_BOUNDS": str(bbs[n]).strip("[]"),
                    "OUTPUT": clipped_vrt,
                },
            )
            # clip_task.run()
            # 'masking layers (part {} of {})'.format(n + 1, len(wkts))
            #
            if self.isCanceled():
                return False
            if not clip_task["SUCCESS"]:
                self.exception = Exception("Clipping failed")
                return False

            ######################################################################
            #  Calculate carbon change table
            logger.log("Calculating summary table...")
            summary_task = processing.run(
                "trendsearth:carbon_summary",
                {
                    "INPUT": clipped_vrt,
                    "year_initial": self.year_initial,
                    "year_final": self.year_final,
                },
            )
            # 'calculating summary table (part {} of {})'.format(n + 1,
            # len(wkts))
            if self.isCanceled():
                return
            if not clip_task["SUCCESS"]:
                self.exception = Exception("Summarizing carbon change failed")
                return False

            os.remove(indic_vrt)
            os.remove(clipped_vrt)
            os.remove(self.tc_vrt)
            os.remove(self.f_loss_vrt)

            if n == 0:
                forest_loss = summary.np_array_from_str(summary_task["FOREST_LOSS"])
                carbon_loss = summary.np_array_from_str(summary_task["CARBON_LOSS"])
                initial_carbon_total = summary_task["CARBON_INITIAL"]
                area_forest = summary_task["AREA_FOREST"]
                area_non_forest = summary_task["AREA_NON_FOREST"]
                area_water = summary_task["AREA_WATER"]
                area_missing = summary_task["AREA_MISSING"]
                area_site = summary_task["AREA_SITE"]
            else:

                forest_loss = forest_loss + summary.np_array_from_str(
                    summary_task["FOREST_LOSS"]
                )
                carbon_loss = carbon_loss + summary.np_array_from_str(
                    summary_task["CARBON_LOSS"]
                )
                area_forest = area_forest + summary_task["AREA_FOREST"]
                area_non_forest = area_non_forest + summary_task["AREA_NON_FOREST"]
                area_water = area_water + summary_task["AREA_WATER"]
                area_missing = area_missing + summary_task["AREA_MISSING"]
                area_site = area_site + summary_task["AREA_SITE"]
                initial_carbon_total = (
                    initial_carbon_total + summary_task["CARBON_INITIAL"]
                )

        logger.log("area_missing: {}".format(area_missing))
        logger.log("area_water: {}".format(area_water))
        logger.log("area_non_forest: {}".format(area_non_forest))
        logger.log("area_site: {}".format(area_site))
        logger.log("area_forest: {}".format(area_forest))
        logger.log("initial_carbon_total: {}".format(initial_carbon_total))
        logger.log("forest loss: {}".format(forest_loss))
        logger.log("carbon loss: {}".format(carbon_loss))

        write_excel_summary(
            forest_loss,
            carbon_loss,
            area_missing,
            area_water,
            area_non_forest,
            area_forest,
            initial_carbon_total,
            self.year_initial,
            self.year_final,
            self.output_file,
        )
        return True

    def finished(self, result):
        if self.isCanceled():
            return
        elif result:
            QtWidgets.QMessageBox.information(
                None,
                tr_calculate_tc.tr("Success"),
                tr_calculate_tc.tr(f"Summary table saved to {self.output_file}"),
            )
        else:
            QtWidgets.QMessageBox.critical(
                None,
                tr_calculate_tc.tr("Error"),
                tr_calculate_tc.tr(
                    f"Error saving output table - check that {self.output_file} is "
                    f"accessible and not already open."
                ),
            )


def write_excel_summary(
    forest_loss,
    carbon_loss,
    area_missing,
    area_water,
    area_non_forest,
    area_forest,
    initial_carbon_total,
    year_initial,
    year_final,
    out_file,
):
    wb = openpyxl.load_workbook(
        Path(__file__).parents[1] / "data" / "summary_table_tc.xlsx"
    )

    ##########################################################################
    # SDG table
    ws_summary = wb["Total Carbon Summary Table"]
    ws_summary.cell(8, 3).value = area_forest
    ws_summary.cell(9, 3).value = area_non_forest
    ws_summary.cell(10, 3).value = area_water
    ws_summary.cell(11, 3).value = area_missing
    ws_summary.cell(15, 4).value = year_initial
    ws_summary.cell(16, 4).value = year_final

    ws_summary.cell(8, 5).value = initial_carbon_total
    summary.write_col_to_sheet(
        ws_summary, np.arange(year_initial + 1, year_final + 1), 1, 24
    )  # Years
    summary.write_col_to_sheet(ws_summary, forest_loss, 2, 24)  # Years
    summary.write_col_to_sheet(ws_summary, carbon_loss, 4, 24)  # Years

    try:
        ws_summary_logo = Image(
            Path(__file__).parents[1] / "data" / "trends_earth_logo_bl_300width.png"
        )
        ws_summary.add_image(ws_summary_logo, "E1")
    except ImportError:
        # add_image will fail on computers without PIL installed (this will be
        # an issue on some Macs, likely others). it is only used here to add
        # our logo, so no big deal.
        pass

    try:
        wb.save(out_file)

    except OSError as exc:
        logger.log(f"Error saving {out_file}: {str(exc)}")
        return False

    logger.log(f"Summary table saved to {out_file}")
    return True
