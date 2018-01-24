# -*- coding: utf-8 -*-
"""
/***************************************************************************
 LDMP - A QGIS plugin
 This plugin supports monitoring and reporting of land degradation to the UNCCD 
 and in support of the SDG Land Degradation Neutrality (LDN) target.
                              -------------------
        begin                : 2017-05-23
        git sha              : $Format:%H$
        copyright            : (C) 2017 by Conservation International
        email                : trends.earth@conservation.org
 ***************************************************************************/
"""

import os
import json
import tempfile

import numpy as np

from osgeo import ogr, osr, gdal

from trendsearth_schemas.schemas import BandInfo, FileList, LocalResults, LocalResultsSchema

import openpyxl
from openpyxl.drawing.image import Image

from PyQt4 import QtGui, uic, QtXml
from PyQt4.QtCore import QSettings, QEventLoop

from qgis.core import QgsGeometry, QgsProject, QgsLayerTreeLayer, QgsLayerTreeGroup, \
    QgsRasterLayer, QgsColorRampShader, QgsRasterShader, \
    QgsSingleBandPseudoColorRenderer, QgsVectorLayer, QgsFeature, \
    QgsCoordinateReferenceSystem, QgsCoordinateTransform, \
    QgsVectorFileWriter, QgsMapLayerRegistry, QgsMapSettings, QgsComposition
from qgis.gui import QgsComposerView
from qgis.utils import iface
mb = iface.messageBar()

from LDMP import log
from LDMP.calculate import DlgCalculateBase
from LDMP.load_data import get_results
from LDMP.plot import DlgPlotBars
from LDMP.gui.DlgReporting import Ui_DlgReporting
from LDMP.gui.DlgReportingSDG import Ui_DlgReportingSDG
from LDMP.gui.DlgCreateMap import Ui_DlgCreateMap
from LDMP.worker import AbstractWorker, start_worker


def create_local_json_metadata(d, outfile, file_format='tif'):
    outfile = os.path.splitext(outfile)[0] + '.json'
    d['results']['local_format'] = file_format
    d['results']['ldmp_version'] = __version__
    with open(outfile, 'w') as outfile:
        json.dump(d, outfile, default=json_serial, sort_keys=True,
                  indent=4, separators=(',', ': '))


def get_band_info(data_file):
    json_file = os.path.splitext(data_file)[0] + '.json'
    res = get_results(json_file)
    if res:
        return res['bands']
    else:
        return None


def _get_layers(node):
    l = []
    if isinstance(node, QgsLayerTreeGroup):
        for child in node.children():
            if isinstance(child, QgsLayerTreeLayer):
                l.append(child.layer())
            else:
                l.extend(_get_layers(child))
    else:
        l = node
    return l


#  Calculate the area of a slice of the globe from the equator to the parallel
#  at latitude f (on WGS84 ellipsoid). Based on:
# https://gis.stackexchange.com/questions/127165/more-accurate-way-to-calculate-area-of-rasters
def _slice_area(f):
    a = 6378137 # in meters
    b = 6356752.3142 # in meters,
    e = np.sqrt(1 - np.square(b / a))
    zp = 1 + e * np.sin(f)
    zm = 1 - e * np.sin(f)
    return np.pi * np.square(b) * ((2 * np.arctanh(e * np.sin(f))) / (2 * e) + np.sin(f) / (zp * zm))


# Formula to calculate area of a raster cell, following
# https://gis.stackexchange.com/questions/127165/more-accurate-way-to-calculate-area-of-rasters
def calc_cell_area(ymin, ymax, x_width):
    'Calculate cell area on WGS84 ellipsoid'
    if ymin > ymax:
        temp = ymax
        ymax = ymin
        ymin = temp
    # ymin: minimum latitude
    # ymax: maximum latitude
    # x_width: width of cell in degrees
    return (_slice_area(np.deg2rad(ymax)) - _slice_area(np.deg2rad(ymin))) * (x_width / 360.)


# Get a list of layers of a particular type, out of those in the TOC that were
# produced by trends.earth
def get_ld_layers(layer_type=None):
    root = QgsProject.instance().layerTreeRoot()
    layers_filtered = []
    layers = _get_layers(root)
    if len(layers) > 0:
        for l in layers:
            if not isinstance(l, QgsRasterLayer):
                # Allows skipping other layer types, like OpenLayers layers, that
                # are irrelevant for the toolbox
                continue
            band_infos = get_band_info(l.dataProvider().dataSourceUri())
            # Layers not produced by trends.earth won't have bandinfo, and 
            # aren't of interest, so skip if there is no bandinfo.
            if band_infos:
                band_number = l.renderer().usesBands()
                # Note the below is true so long as none of the needed layers use more 
                # than one band.
                if len(band_number) == 1:
                    band_number = band_number[0]
                    band_info = [band_info for band_info in band_infos if band_info['band_number'] == band_number][0]
                    name = band_info['name']
                    if layer_type == 'traj_sig' and name == 'Productivity trajectory (significance)':
                        layers_filtered.append((l, band_number, band_info))
                    elif layer_type == 'state_deg' and name == 'Productivity state (degradation)':
                        layers_filtered.append((l, band_number, band_info))
                    elif layer_type == 'perf_deg' and name == 'Productivity performance (degradation)':
                        layers_filtered.append((l, band_number, band_info))
                    elif layer_type == 'lc_tr' and name == 'Land cover transitions':
                        layers_filtered.append((l, band_number, band_info))
                    elif layer_type == 'lc_deg' and name == 'Land cover degradation':
                        layers_filtered.append((l, band_number, band_info))
                    elif layer_type == 'soc_deg' and name == 'Soil organic carbon (degradation)':
                        layers_filtered.append((l, band_number, band_info))
                    elif layer_type == 'soc_annual' and name == 'Soil organic carbon':
                        layers_filtered.append((l, band_number, band_info))
                    elif layer_type == 'lc_mode' and name == 'Land cover mode (7 class)':
                        layers_filtered.append((l, band_number, band_info))
                    elif layer_type == 'lc_annual' and name == 'Land cover (7 class)':
                        layers_filtered.append((l, band_number, band_info))
                    elif layer_type == 'lc_transitions' and name == 'Land cover transitions':
                        layers_filtered.append((l, band_number, band_info))
    return layers_filtered


def style_sdg_ld(outfile, title):
    # Significance layer
    log('Loading layers onto map.')
    layer = iface.addRasterLayer(outfile, title)
    if not layer.isValid():
        log('Failed to add layer')
        return None
    fcn = QgsColorRampShader()
    fcn.setColorRampType(QgsColorRampShader.EXACT)
    lst = [QgsColorRampShader.ColorRampItem(-32768, QtGui.QColor(0, 0, 0), QtGui.QApplication.translate('LDMPPlugin', 'No data')),
           QgsColorRampShader.ColorRampItem(-32767, QtGui.QColor(190, 190, 190), QtGui.QApplication.translate('LDMPPlugin', 'Masked area')),
           QgsColorRampShader.ColorRampItem(-1, QtGui.QColor(153, 51, 4), QtGui.QApplication.translate('LDMPPlugin', 'Degradation')),
           QgsColorRampShader.ColorRampItem(0, QtGui.QColor(246, 246, 234), QtGui.QApplication.translate('LDMPPlugin', 'Stable')),
           QgsColorRampShader.ColorRampItem(1, QtGui.QColor(0, 140, 121), QtGui.QApplication.translate('LDMPPlugin', 'Improvement'))]
           
    fcn.setColorRampItemList(lst)
    shader = QgsRasterShader()
    shader.setRasterShaderFunction(fcn)
    pseudoRenderer = QgsSingleBandPseudoColorRenderer(layer.dataProvider(), 1, shader)
    layer.setRenderer(pseudoRenderer)
    layer.triggerRepaint()
    iface.legendInterface().refreshLayerSymbology(layer)


class DegradationWorkerSDG(AbstractWorker):
    def __init__(self, src_file, deg_file, prod_file):
        AbstractWorker.__init__(self)

        self.src_file = src_file
        self.deg_file = deg_file
        self.prod_file = prod_file

    def work(self):
        self.toggle_show_progress.emit(True)
        self.toggle_show_cancel.emit(True)

        src_ds = gdal.Open(self.src_file)

        traj_band = src_ds.GetRasterBand(1)
        perf_band = src_ds.GetRasterBand(2)
        state_band = src_ds.GetRasterBand(3)
        lc_band = src_ds.GetRasterBand(4)
        soc_band = src_ds.GetRasterBand(5)

        block_sizes = traj_band.GetBlockSize()
        x_block_size = block_sizes[0]
        y_block_size = block_sizes[1]
        xsize = traj_band.XSize
        ysize = traj_band.YSize

        driver = gdal.GetDriverByName("GTiff")
        dst_ds_deg = driver.Create(self.deg_file, xsize, ysize, 1, gdal.GDT_Int16, ['COMPRESS=LZW'])
        # Save the combined productivity indicator as well
        dst_ds_prod = driver.Create(self.prod_file, xsize, ysize, 1, gdal.GDT_Int16, ['COMPRESS=LZW'])

        src_gt = src_ds.GetGeoTransform()
        dst_ds_deg.SetGeoTransform(src_gt)
        dst_ds_prod.SetGeoTransform(src_gt)
        dst_srs = osr.SpatialReference()
        dst_srs.ImportFromWkt(src_ds.GetProjectionRef())
        dst_ds_deg.SetProjection(dst_srs.ExportToWkt())
        dst_ds_prod.SetProjection(dst_srs.ExportToWkt())

        xsize = traj_band.XSize
        ysize = traj_band.YSize
        blocks = 0
        for y in xrange(0, ysize, y_block_size):
            if self.killed:
                log("Processing of {} killed by user after processing {} out of {} blocks.".format(deg_file, y, ysize))
                break
            self.progress.emit(100 * float(y) / ysize)
            if y + y_block_size < ysize:
                rows = y_block_size
            else:
                rows = ysize - y
            for x in xrange(0, xsize, x_block_size):
                if x + x_block_size < xsize:
                    cols = x_block_size
                else:
                    cols = xsize - x

                # TODO: Could make this cleaner by reading all four bands at
                # same time from VRT
                traj_array = traj_band.ReadAsArray(x, y, cols, rows)
                state_array = state_band.ReadAsArray(x, y, cols, rows)
                perf_array = perf_band.ReadAsArray(x, y, cols, rows)
                lc_array = lc_band.ReadAsArray(x, y, cols, rows)
                soc_array = soc_band.ReadAsArray(x, y, cols, rows)

                ##############
                # Productivity
                
                # Capture trends that are at least 95% significant.
                # Remember that traj is coded as:
                # -3: 99% signif decline
                # -2: 95% signif decline
                # -1: 90% signif decline
                #  0: stable
                #  1: 90% signif increase
                #  2: 95% signif increase
                #  3: 99% signif increase
                deg = traj_array
                deg[deg == -1] = 0 # not signif at 95%
                deg[deg == 1] = 0 # not signif at 95%
                deg[np.logical_and(deg >= -3, deg <= -2)] = -1
                deg[np.logical_and(deg >= 2, deg <= 3)] = 1

                # Handle state and performance. Note that state array is the 
                # number of changes in class, so  <= -2 is a decline.
                deg[np.logical_and(np.logical_and(state_array <= -2, state_array >= -10), perf_array == -1)] = -1

                # Ensure NAs carry over to productivity indicator layer
                deg[traj_array == -32768] = -32768
                deg[perf_array == -32768] = -32768
                deg[state_array == -32768] = -32768
                # Ensure masked areas carry over to productivity indicator 
                # layer
                deg[traj_array == -32767] = -32767
                deg[perf_array == -32767] = -32767
                deg[state_array == -32767] = -32767

                # Save combined productivity indicator for later visualization
                dst_ds_prod.GetRasterBand(1).WriteArray(deg, x, y)

                #############
                # Land cover
                deg[lc_array == -1] = -1

                ##############
                # Soil carbon
                
                # Note SOC array is coded in percent change, so change of 
                # greater than 10% is improvement or decline.
                deg[np.logical_and(soc_array <= -10, soc_array >= -100)] = -1

                #############
                # Improvement
                
                # Allow improvements by lc or soc, only where one of the other 
                # two indicators doesn't indicate a decline
                deg[np.logical_and(deg == 0, lc_array == 1)] = 1
                deg[np.logical_and(deg == 0, np.logical_and(soc_array >= 10, soc_array <= 100))] = 1

                ##############
                # Missing data
                
                # Ensure all NAs are carried over - note this was already done 
                # for the productivity layer above but need to do it again in 
                # case values from another layer overwrote those missing value 
                # indicators.
                
                # No data
                deg[traj_array == -32768] = -32768
                deg[perf_array == -32768] = -32768
                deg[state_array == -32768] = -32768
                deg[lc_array == -32768] = -32768
                deg[soc_array == -32768] = -32768

                # Masked areas
                deg[traj_array == -32767] = -32767
                deg[perf_array == -32767] = -32767
                deg[state_array == -32767] = -32767
                deg[lc_array == -32767] = -32767
                deg[soc_array == -32767] = -32767

                dst_ds_deg.GetRasterBand(1).WriteArray(deg, x, y)
                del deg
                blocks += 1
        self.progress.emit(100)
        src_ds = None
        dst_ds = None

        if self.killed:
            os.remove(deg_file)
            os.remove(prod_file)
            return None
        else:
            return True


def xtab(*cols):
    # Based on https://gist.github.com/alexland/d6d64d3f634895b9dc8e, but
    # modified to ignore np.nan
    if not all(len(col) == len(cols[0]) for col in cols[1:]):
        raise ValueError("all arguments must be same size")

    if len(cols) == 0:
        raise TypeError("xtab() requires at least one argument")

    def fnx1(q): return len(q.squeeze().shape)
    if not all([fnx1(col) == 1 for col in cols]):
        raise ValueError("all input arrays must be 1D")

    # Filter na values out of all columns
    nafilter = ~np.any(np.isnan(cols), 0)

    headers, idx = zip(*(np.unique(col[nafilter], return_inverse=True) for col in cols))
    shape_xt = [uniq_vals_col.size for uniq_vals_col in headers]
    xt = np.zeros(shape_xt)
    np.add.at(xt, idx, 1)

    return list((headers, xt))


def merge_xtabs(tab1, tab2):
    """Mergies two crosstabs - allows for block-by-block crosstabs"""
    headers = tuple(np.array(np.unique(np.concatenate(header))) for header in zip(tab1[0], tab2[0]))
    shape_xt = [uniq_vals_col.size for uniq_vals_col in headers]
    # Make this array flat since it will be used later with ravelled indexing
    xt = np.zeros(np.prod(shape_xt))

    # This handles combining a crosstab from a new block with an existing one
    # that has been maintained across blocks
    def add_xt_block(xt_bl):
        col_ind = np.tile(tuple(np.where(headers[0] == item) for item in xt_bl[0][0]), xt_bl[0][1].size)
        row_ind = np.transpose(np.tile(tuple(np.where(headers[1] == item) for item in xt_bl[0][1]), xt_bl[0][0].size))
        ind = np.ravel_multi_index((col_ind, row_ind), shape_xt)
        np.add.at(xt, ind.ravel(), xt_bl[1].ravel())
    add_xt_block(tab1)
    add_xt_block(tab2)

    return list((headers, xt.reshape(shape_xt)))


def calc_total_table(a_trans, a_soc, total_table, cell_area):
    """Calculates an total table for an array"""
    if total_table:
        # Add in totals for past total_table if one is provided
        transitions = np.unique(np.concatenate([a_trans.ravel(), total_table[0]]))
        ind = np.concatenate(tuple(np.where(transitions == item)[0] for item in total_table[0]))
        totals = np.zeros(transitions.shape)
        np.add.at(totals, ind, total_table[1])
    else:
        transitions = np.unique(np.concatenate(a_trans))
        totals = np.zeros(transitions.shape)

    for transition in transitions:
        ind = np.where(transitions == transition)
        # Only sum values for this transition, and where soc has a valid value
        # (negative values are missing data flags)
        vals = a_soc[np.logical_and(a_trans == transition, a_soc > 0)]
        totals[ind] += np.sum(vals * cell_area)

    return list((transitions, totals))


def calc_area_table(a, area_table, cell_area):
    """Calculates an area table for an array"""
    # Convert array to int32 dtype so that the correction below can be applied 
    # without the array overflowing
    a = a.astype(np.int32)
    a_min = np.min(a)
    if a_min < 0:
        # Correction to add as bincount can only handle positive integers
        correction = np.abs(a_min)
    else:
        correction = 0

    n = np.bincount(a.ravel() + correction)
    this_vals = np.nonzero(n)[0]
    # Subtract correction from this_vals so area table has correct values
    this_area_table = list([this_vals - correction, n[this_vals]])

    # Don't use this_area_table if it is empty
    if this_area_table[0].size != 0:
        this_area_table[1] = this_area_table[1] * cell_area
        if area_table == None:
            area_table = this_area_table
        else:
            area_table = merge_area_tables(area_table, this_area_table)
    return area_table


def merge_area_tables(table1, table2):
    vals = np.unique(np.concatenate([table1[0], table2[0]]))
    count = np.zeros(vals.shape)

    def add_area_table(table):
        ind = np.concatenate(tuple(np.where(vals == item)[0] for item in table[0]))
        np.add.at(count, ind, table[1])
    add_area_table(table1)
    add_area_table(table2)
    return list((vals, count))


class AreaWorker(AbstractWorker):
    def __init__(self, in_file):
        AbstractWorker.__init__(self)
        self.in_file = in_file

    def work(self):
        ds = gdal.Open(self.in_file)
        band_deg = ds.GetRasterBand(1)
        band_lc_bl = ds.GetRasterBand(2)
        band_lc_tg = ds.GetRasterBand(3)
        band_soc_bl = ds.GetRasterBand(4)
        band_soc_tg = ds.GetRasterBand(5)

        block_sizes = band_deg.GetBlockSize()
        x_block_size = block_sizes[0]
        # Need to process y line by line so that pixel area calculation can be
        # done based on latitude, which varies by line
        y_block_size = 1
        xsize = band_deg.XSize
        ysize = band_deg.YSize

        gt = ds.GetGeoTransform()
        # Width of cells in longitude
        long_width = gt[1]

        # Set initial lat ot the top left corner latitude
        lat = gt[3]
        # Width of cells in latitude
        pixel_height = gt[5]

        blocks = 0
        trans_xtab = None
        soc_bl_totals_table = None
        soc_tg_totals_table = None
        for y in xrange(0, ysize, y_block_size):
            if self.killed:
                log("Processing killed by user after processing {} out of {} blocks.".format(y, ysize))
                break
            self.progress.emit(100 * float(y) / ysize)
            if y + y_block_size < ysize:
                rows = y_block_size
            else:
                rows = ysize - y
            for x in xrange(0, xsize, x_block_size):
                if x + x_block_size < xsize:
                    cols = x_block_size
                else:
                    cols = xsize - x

                cell_area = calc_cell_area(lat, lat + pixel_height, long_width)

                ################################
                # Calculate transition crosstabs
                a_lc_bl = band_lc_bl.ReadAsArray(x, y, cols, rows)
                a_lc_tg = band_lc_tg.ReadAsArray(x, y, cols, rows)
                a_trans = a_lc_bl*10 + a_lc_tg
                a_trans[np.logical_or(a_lc_bl < 1, a_lc_tg < 1)] <- -32768

                a_deg = band_deg.ReadAsArray(x, y, cols, rows)

                # Flatten the arrays before passing to xtab
                this_trans_xtab = xtab(a_deg.ravel(), a_trans.ravel())

                # Don't use this_trans_xtab if it is empty (could happen if take a
                # crosstab where all of the values are nan's)
                if this_trans_xtab[0][0].size != 0:
                    this_trans_xtab[1] = this_trans_xtab[1] * cell_area
                    if trans_xtab == None:
                        trans_xtab = this_trans_xtab
                    else:
                        trans_xtab = merge_xtabs(trans_xtab, this_trans_xtab)

                #################################
                # Calculate SOC totals (converting soilgrids data from per ha
                # to per meter since cell_area is in meters). Note final units 
                # of soc_totals tables are tons C (summed over the total area 
                # of each class)
                a_soc_bl = band_soc_bl.ReadAsArray(x, y, cols, rows).astype(np.float32) / (100 * 100)
                soc_bl_totals_table = calc_total_table(a_trans, a_soc_bl,
                                                       soc_bl_totals_table, cell_area)
                a_soc_tg = band_soc_tg.ReadAsArray(x, y, cols, rows).astype(np.float32) / (100 * 100)
                soc_tg_totals_table = calc_total_table(a_trans, a_soc_tg,
                                                       soc_tg_totals_table, cell_area)

                blocks += 1
            lat += pixel_height
        self.progress.emit(100)
        self.ds = None

        # Convert all area tables from meters into square kilometers
        trans_xtab[1] = trans_xtab[1] * 1e-6

        if self.killed:
            return None
        else:
            return list((soc_bl_totals_table, soc_tg_totals_table, trans_xtab))


# Returns value from crosstab table for particular deg/lc class combination
def get_xtab_area(table, deg_class=None, lc_class=None):
    deg_ind = np.where(table[0][0] == deg_class)[0]
    lc_ind = np.where(table[0][1] == lc_class)[0]
    if deg_ind.size != 0 and lc_ind.size != 0:
        return float(table[1][deg_ind, lc_ind])
    elif deg_ind.size != 0 and lc_class == None:
        return float(np.sum(table[1][deg_ind, :]))
    elif lc_ind.size != 0 and deg_class == None:
        return float(np.sum(table[1][:, lc_ind]))
    elif lc_class == None and deg_class == None:
        return float(np.sum(table[1].ravel()))
    else:
        return 0


class ClipWorker(AbstractWorker):
    def __init__(self, in_file, out_file, mask_layer):
        AbstractWorker.__init__(self)

        self.in_file = in_file
        self.out_file = out_file

        self.mask_layer = mask_layer

    def work(self):
        self.toggle_show_progress.emit(True)
        self.toggle_show_cancel.emit(True)

        mask_layer_file = tempfile.NamedTemporaryFile(suffix='.shp').name
        QgsVectorFileWriter.writeAsVectorFormat(self.mask_layer, mask_layer_file,
                                                "CP1250", None, "ESRI Shapefile")

        res = gdal.Warp(self.out_file, self.in_file, format='GTiff',
                        cutlineDSName=mask_layer_file,
                        srcNodata=-32768, dstNodata=-32767,
                        dstSRS="epsg:4326",
                        outputType=gdal.GDT_Int16,
                        resampleAlg=gdal.GRA_NearestNeighbour,
                        creationOptions=['COMPRESS=LZW'],
                        callback=self.progress_callback)

        if res:
            return True
        else:
            return None

    def progress_callback(self, fraction, message, data):
        if self.killed:
            return False
        else:
            self.progress.emit(100 * fraction)
            return True


class StartWorker(object):
    def __init__(self, worker_class, process_name, *args):
        self.exception = None
        self.success = None

        self.worker = worker_class(*args)

        pause = QEventLoop()
        self.worker.finished.connect(pause.quit)
        self.worker.successfully_finished.connect(self.save_success)
        self.worker.error.connect(self.save_exception)
        start_worker(self.worker, iface,
                     QtGui.QApplication.translate("LDMP", 'Processing: {}').format(process_name))
        pause.exec_()

        if self.exception:
            raise self.exception

    def save_success(self, val=None):
        self.return_val = val
        self.success = True

    def get_return(self):
        return self.return_val

    def save_exception(self, exception):
        self.exception = exception

    def get_exception(self):
        return self.exception


class DlgReporting(QtGui.QDialog, Ui_DlgReporting):
    def __init__(self, parent=None):
        super(DlgReporting, self).__init__(parent)
        self.setupUi(self)

        self.dlg_sdg = DlgReportingSDG()
        self.dlg_create_map = DlgCreateMap()

        self.btn_sdg.clicked.connect(self.clicked_sdg)
        self.btn_create_map.clicked.connect(self.clicked_create_map)

    def clicked_create_map(self):
        self.close()
        self.dlg_create_map.exec_()

    def clicked_sdg(self):
        self.close()
        self.dlg_sdg.exec_()


class DlgReportingSDG(DlgCalculateBase, Ui_DlgReportingSDG):
    '''Class to be shared across SDG and SummaryTable reporting dialogs'''

    def __init__(self, parent=None):
        super(DlgReportingSDG, self).__init__(parent)
        self.setupUi(self)

        self.mode_lpd.toggled.connect(self.mode_lpd_toggled)

    def mode_lpd_toggled(self):
        if self.mode_lpd.isChecked():
            self.combo_layer_lpd.setEnabled(True)
            self.combo_layer_traj.setEnabled(False)
            self.combo_layer_traj_label.setEnabled(False)
            self.combo_layer_perf.setEnabled(False)
            self.combo_layer_perf_label.setEnabled(False)
            self.combo_layer_state.setEnabled(False)
            self.combo_layer_state_label.setEnabled(False)
        else:
            self.combo_layer_lpd.setEnabled(False)
            self.combo_layer_traj.setEnabled(True)
            self.combo_layer_traj_label.setEnabled(True)
            self.combo_layer_perf.setEnabled(True)
            self.combo_layer_perf_label.setEnabled(True)
            self.combo_layer_state.setEnabled(True)
            self.combo_layer_state_label.setEnabled(True)

    def showEvent(self, event):
        super(DlgReportingSDG, self).showEvent(event)
        #self.populate_layers_lpd()
        self.populate_layers_traj()
        self.populate_layers_perf()
        self.populate_layers_state()
        self.populate_layers_lc()
        self.populate_layers_soc()

    def firstShow(self):
        self.browse_output_file_layer.clicked.connect(self.select_output_file_layer)
        self.browse_output_file_table.clicked.connect(self.select_output_file_table)
        super(DlgReportingSDG, self).firstShow()

    #def populate_layers_lpd(self):
    #    self.combo_layer_lpd.clear()
    #    self.layer_lpd_list = get_ld_layers('lpd')
    #    self.combo_layer_lpd.addItems([l[0].name() for l in self.layer_lpd_list])

    def populate_layers_traj(self):
        self.combo_layer_traj.clear()
        self.layer_traj_list = get_ld_layers('traj_sig')
        self.combo_layer_traj.addItems([l[0].name() for l in self.layer_traj_list])

    def populate_layers_perf(self):
        self.combo_layer_perf.clear()
        self.layer_perf_list = get_ld_layers('perf_deg')
        self.combo_layer_perf.addItems([l[0].name() for l in self.layer_perf_list])

    def populate_layers_state(self):
        self.combo_layer_state.clear()
        self.layer_state_list = get_ld_layers('state_deg')
        self.combo_layer_state.addItems([l[0].name() for l in self.layer_state_list])

    def populate_layers_lc(self):
        self.combo_layer_lc.clear()
        self.layer_lc_list = get_ld_layers('lc_deg')
        self.combo_layer_lc.addItems([l[0].name() for l in self.layer_lc_list])

    def populate_layers_soc(self):
        self.combo_layer_soc.clear()
        self.layer_soc_list = get_ld_layers('soc_deg')
        self.combo_layer_soc.addItems([l[0].name() for l in self.layer_soc_list])

    def select_output_file_layer(self):
        f = QtGui.QFileDialog.getSaveFileName(self,
                                              self.tr('Choose a filename for the output file'),
                                              QSettings().value("LDMP/output_dir", None),
                                              self.tr('GeoTIFF file (*.tif)'))
        if f:
            if os.access(os.path.dirname(f), os.W_OK):
                QSettings().setValue("LDMP/output_dir", os.path.dirname(f))
                self.output_file_layer.setText(f)
            else:
                QtGui.QMessageBox.critical(None, self.tr("Error"),
                                           self.tr("Cannot write to {}. Choose a different file.".format(f), None))

    def select_output_file_table(self):
        f = QtGui.QFileDialog.getSaveFileName(self,
                                              self.tr('Choose a filename for the summary table'),
                                              QSettings().value("LDMP/output_dir", None),
                                              self.tr('Summary table file (*.xlsx)'))
        if f:
            if os.access(os.path.dirname(f), os.W_OK):
                QSettings().setValue("LDMP/output_dir", os.path.dirname(f))
                self.output_file_table.setText(f)
            else:
                QtGui.QMessageBox.critical(None, self.tr("Error"),
                                           self.tr("Cannot write to {}. Choose a different file.".format(f), None))

    def get_resample_alg(self, lc_f, traj_f):
        ds_lc = gdal.Open(lc_f)
        ds_traj = gdal.Open(traj_f)
        # If prod layers are lower res than the lc layer, then resample lc
        # using the mode. Otherwise use nearest neighbor:
        lc_gt = ds_lc.GetGeoTransform()
        traj_gt = ds_traj.GetGeoTransform()
        if lc_gt[1] < traj_gt[1]:
            # If the land cover is finer than the trajectory res, use mode to
            # match the lc to the lower res productivity data
            log('Resampling with: mode, lowest')
            return('lowest', gdal.GRA_Mode)
        else:
            # If the land cover is coarser than the trajectory res, use nearest
            # neighbor and match the lc to the higher res productivity data
            log('Resampling with: nearest neighour, highest')
            return('highest', gdal.GRA_NearestNeighbour)

    def btn_calculate(self):
        ######################################################################
        # Check that all needed output files are selected
        if not self.output_file_layer.text():
            QtGui.QMessageBox.information(None, self.tr("Error"),
                                          self.tr("Choose an output file for the indicator layer."), None)
            return

        if not self.output_file_table.text():
            QtGui.QMessageBox.information(None, self.tr("Error"),
                                          self.tr("Choose an output file for the summary table."), None)
            return

        # Note that the super class has several tests in it - if they fail it
        # returns False, which would mean this function should stop execution
        # as well.
        ret = super(DlgReportingSDG, self).btn_calculate()
        if not ret:
            return

        ######################################################################
        # Check that all needed input layers are selected
        if len(self.layer_traj_list) == 0:
            QtGui.QMessageBox.critical(None, self.tr("Error"),
                                       self.tr("You must add a productivity trajectory indicator layer to your map before you can use the reporting tool."), None)
            return
        if len(self.layer_state_list) == 0:
            QtGui.QMessageBox.critical(None, self.tr("Error"),
                                       self.tr("You must add a productivity state indicator layer to your map before you can use the reporting tool."), None)
            return
        if len(self.layer_perf_list) == 0:
            QtGui.QMessageBox.critical(None, self.tr("Error"),
                                       self.tr("You must add a productivity performance indicator layer to your map before you can use the reporting tool."), None)
            return
        if len(self.layer_lc_list) == 0:
            QtGui.QMessageBox.critical(None, self.tr("Error"),
                                       self.tr("You must add a productivity land cover indicator layer to your map before you can use the reporting tool."), None)
            return
        if len(self.layer_soc_list) == 0:
            QtGui.QMessageBox.critical(None, self.tr("Error"),
                                       self.tr("You must add a soil organic carbon indicator layer to your map before you can use the reporting tool."), None)
            return

        self.layer_traj = self.layer_traj_list[self.combo_layer_traj.currentIndex()][0]
        self.layer_traj_bandnumber = self.layer_traj_list[self.combo_layer_traj.currentIndex()][1]

        self.layer_perf = self.layer_perf_list[self.combo_layer_perf.currentIndex()][0]
        self.layer_perf_bandnumber = self.layer_perf_list[self.combo_layer_perf.currentIndex()][1]

        self.layer_state = self.layer_state_list[self.combo_layer_state.currentIndex()][0]
        self.layer_state_bandnumber = self.layer_state_list[self.combo_layer_state.currentIndex()][1]

        self.layer_lc = self.layer_lc_list[self.combo_layer_lc.currentIndex()][0]
        self.layer_lc_bandnumber = self.layer_lc_list[self.combo_layer_lc.currentIndex()][1]

        self.layer_soc = self.layer_soc_list[self.combo_layer_soc.currentIndex()][0]
        self.layer_soc_bandnumber = self.layer_soc_list[self.combo_layer_soc.currentIndex()][1]

        #######################################################################
        # Check that the layers cover the full extent needed
        if not self.aoi.bounding_box_geom.within(QgsGeometry.fromRect(self.layer_traj.extent())):
            QtGui.QMessageBox.critical(None, self.tr("Error"),
                                       self.tr("Area of interest is not entirely within the trajectory layer."), None)
            return
        if not self.aoi.bounding_box_geom.within(QgsGeometry.fromRect(self.layer_perf.extent())):
            QtGui.QMessageBox.critical(None, self.tr("Error"),
                                       self.tr("Area of interest is not entirely within the performance layer."), None)
            return
        if not self.aoi.bounding_box_geom.within(QgsGeometry.fromRect(self.layer_state.extent())):
            QtGui.QMessageBox.critical(None, self.tr("Error"),
                                       self.tr("Area of interest is not entirely within the state layer."), None)
            return
        if not self.aoi.bounding_box_geom.within(QgsGeometry.fromRect(self.layer_lc.extent())):
            QtGui.QMessageBox.critical(None, self.tr("Error"),
                                       self.tr("Area of interest is not entirely within the land cover layer."), None)
            return
        if not self.aoi.bounding_box_geom.within(QgsGeometry.fromRect(self.layer_soc.extent())):
            QtGui.QMessageBox.critical(None, self.tr("Error"),
                                       self.tr("Area of interest is not entirely within the soil organic carbon layer."), None)
            return

        #######################################################################
        # Check that all of the productivity layers have the same resolution
        def res(layer):
            return (round(layer.rasterUnitsPerPixelX(), 10), round(layer.rasterUnitsPerPixelY(), 10))
        if res(self.layer_traj) != res(self.layer_state):
            QtGui.QMessageBox.critical(None, self.tr("Error"),
                                       self.tr("Resolutions of trajectory layer and state layer do not match."), None)
            return
        if res(self.layer_traj) != res(self.layer_perf):
            QtGui.QMessageBox.critical(None, self.tr("Error"),
                                       self.tr("Resolutions of trajectory layer and performance layer do not match."), None)
            return

        #######################################################################
        # Check that all of the productivity layers have the same CRS
        if self.layer_traj.crs() != self.layer_state.crs():
            QtGui.QMessageBox.critical(None, self.tr("Error"),
                                       self.tr("Coordinate systems of trajectory layer and state layer do not match."), None)
            return
        if self.layer_traj.crs() != self.layer_perf.crs():
            QtGui.QMessageBox.critical(None, self.tr("Error"),
                                       self.tr("Coordinate systems of trajectory layer and performance layer do not match."), None)
            return

        self.close()

        #######################################################################
        # Load all datasets to VRTs (to select only the needed bands)
        self.traj_f = tempfile.NamedTemporaryFile(suffix='.vrt').name
        gdal.BuildVRT(self.traj_f, self.layer_traj.dataProvider().dataSourceUri(),
                      bandList=[self.layer_traj_bandnumber])

        self.perf_f = tempfile.NamedTemporaryFile(suffix='.vrt').name
        gdal.BuildVRT(self.perf_f, self.layer_perf.dataProvider().dataSourceUri(),
                      bandList=[self.layer_perf_bandnumber])

        self.state_f = tempfile.NamedTemporaryFile(suffix='.vrt').name
        gdal.BuildVRT(self.state_f, self.layer_state.dataProvider().dataSourceUri(),
                      bandList=[self.layer_state_bandnumber])

        self.lc_deg_f = tempfile.NamedTemporaryFile(suffix='.vrt').name
        gdal.BuildVRT(self.lc_deg_f, self.layer_lc.dataProvider().dataSourceUri(),
                bandList=[self.layer_lc_bandnumber])

        self.soc_deg_f = tempfile.NamedTemporaryFile(suffix='.vrt').name
        gdal.BuildVRT(self.soc_deg_f, self.layer_soc.dataProvider().dataSourceUri(),
                      bandList=[self.layer_soc_bandnumber])

        # Compute the pixel-aligned bounding box (slightly larger than aoi).
        # Use this instead of croptocutline in gdal.Warp in order to keep the
        # pixels aligned.
        bb = self.aoi.bounding_box_geom.boundingBox()
        minx = bb.xMinimum()
        miny = bb.yMinimum()
        maxx = bb.xMaximum()
        maxy = bb.yMaximum()
        traj_gt = gdal.Open(self.traj_f).GetGeoTransform()
        left = minx - (minx - traj_gt[0]) % traj_gt[1]
        right = maxx + (traj_gt[1] - ((maxx - traj_gt[0]) % traj_gt[1]))
        bottom = miny + (traj_gt[5] - ((miny - traj_gt[3]) % traj_gt[5]))
        top = maxy - (maxy - traj_gt[3]) % traj_gt[5]
        self.outputBounds = [left, bottom, right, top]

        #######################################################################
        #######################################################################
        # Calculate SDG 15.3 layers
        #######################################################################
        #######################################################################

        ######################################################################
        # Combine input rasters for SDG 15.3.1 into a VRT and crop to the AOI
        self.indic_f = tempfile.NamedTemporaryFile(suffix='.vrt').name
        log('Saving indicator VRT to: {}'.format(self.indic_f))
        resample_alg = self.get_resample_alg(self.lc_deg_f, self.traj_f)
        gdal.BuildVRT(self.indic_f,
                      [self.traj_f,
                       self.perf_f,
                       self.state_f,
                       self.lc_deg_f,
                       self.soc_deg_f],
                      outputBounds=self.outputBounds,
                      resolution=resample_alg[0],
                      resampleAlg=resample_alg[1],
                      separate=True)

        lc_clip_tempfile = tempfile.NamedTemporaryFile(suffix='.tif').name
        log('Saving deg/lc clipped file to {}'.format(lc_clip_tempfile))
        deg_lc_clip_worker = StartWorker(ClipWorker, 'masking layers',
                                         self.indic_f,
                                         lc_clip_tempfile, self.aoi.layer)
        if not deg_lc_clip_worker.success:
            QtGui.QMessageBox.critical(None, self.tr("Error"),
                                       self.tr("Error masking SDG 15.3.1 input layers."), None)
            return

        ######################################################################
        #  Calculate SDG 15.3.1 layers
        log('Calculating degradation...')
        prod_f = os.path.splitext(self.output_file_layer.text())[0] + '_Productivity_Sub-Indicator.tif'
        deg_worker = StartWorker(DegradationWorkerSDG, 'calculating degradation',
                                 lc_clip_tempfile, self.output_file_layer.text(), prod_f)
        if not deg_worker.success:
            QtGui.QMessageBox.critical(None, self.tr("Error"),
                                       self.tr("Error calculating SDG 15.3.1 degradation layer."), None)
            return

        #######################################################################
        #######################################################################
        # Produce summary table
        #######################################################################
        #######################################################################

        #######################################################################
        # Select baseline and target land cover and SOC layers based on chosen
        # degradation layers for these datasets

        lc_band_infos = get_band_info(self.layer_lc.dataProvider().dataSourceUri())
        #TODO: Need to fix this code for when drop the mode on land cover calculation
        self.layer_lc_bl_bandnumber = [band_info['band_number'] for band_info in lc_band_infos if band_info['name'] == 'Land cover mode (7 class)'][0]
        self.layer_lc_tg_bandnumber = [band_info['band_number'] for band_info in lc_band_infos if band_info['name'] == 'Land cover (7 class)'][0]

        soc_band_infos = [band_info for band_info in get_band_info(self.layer_soc.dataProvider().dataSourceUri()) if band_info['name'] == 'Soil organic carbon']
        first_year = min([band_info['metadata']['year'] for band_info in soc_band_infos])
        last_year = max([band_info['metadata']['year'] for band_info in soc_band_infos])
        self.layer_soc_bl_bandnumber = [band_info['band_number'] for band_info in soc_band_infos if band_info['metadata']['year'] == first_year][0]
        self.layer_soc_tg_bandnumber = [band_info['band_number'] for band_info in soc_band_infos if band_info['metadata']['year'] == last_year][0]

        ######################################################################
        # Combine rasters into a VRT and crop to the AOI
        indic_f = tempfile.NamedTemporaryFile(suffix='.vrt').name
        log('Saving deg/lc/soc VRT to: {}'.format(indic_f))

        # Select lc bands using bandlist since BuildVrt will otherwise only use
        # the first band of the file
        self.lc_bl_f = tempfile.NamedTemporaryFile(suffix='.vrt').name
        gdal.BuildVRT(self.lc_bl_f, self.layer_lc.dataProvider().dataSourceUri(),
                      bandList=[self.layer_lc_bl_bandnumber])

        self.lc_tg_f = tempfile.NamedTemporaryFile(suffix='.vrt').name
        gdal.BuildVRT(self.lc_tg_f, self.layer_lc.dataProvider().dataSourceUri(),
                      bandList=[self.layer_lc_tg_bandnumber])

        self.soc_bl_f = tempfile.NamedTemporaryFile(suffix='.vrt').name
        gdal.BuildVRT(self.soc_bl_f, self.layer_soc.dataProvider().dataSourceUri(),
                      bandList=[self.layer_soc_bl_bandnumber])

        self.soc_tg_f = tempfile.NamedTemporaryFile(suffix='.vrt').name
        gdal.BuildVRT(self.soc_tg_f, self.layer_soc.dataProvider().dataSourceUri(),
                      bandList=[self.layer_soc_tg_bandnumber])

        gdal.BuildVRT(indic_f,
                      [prod_f,
                       self.lc_bl_f,
                       self.lc_tg_f,
                       self.soc_bl_f,
                       self.soc_tg_f],
                      outputBounds=self.outputBounds,
                      resolution=resample_alg[0],
                      resampleAlg=resample_alg[1],
                      separate=True)
        # Clip and mask the lc/deg layer before calculating crosstab
        lc_clip_tempfile = tempfile.NamedTemporaryFile(suffix='.tif').name
        log('Saving deg/lc clipped file to {}'.format(lc_clip_tempfile))
        deg_lc_clip_worker = StartWorker(ClipWorker, 'masking layers',
                                         indic_f,
                                         lc_clip_tempfile, self.aoi.layer)
        if not deg_lc_clip_worker.success:
            QtGui.QMessageBox.critical(None, self.tr("Error"),
                                       self.tr("Error clipping layers for area calculation."), None)
            return

        ######################################################################
        # Calculate area crosstabs

        log('Calculating land cover crosstabulation...')
        area_worker = StartWorker(AreaWorker, 'calculating areas', lc_clip_tempfile)
        if not area_worker.success:
            QtGui.QMessageBox.critical(None, self.tr("Error"),
                                       self.tr("Error calculating degraded areas."), None)
            return
        else:
            soc_bl_totals, soc_tg_totals, trans_prod_xtab = area_worker.get_return()
            log('Soc bl: {}'.format(soc_bl_totals))

        x = [self.tr('Degraded'), self.tr('Stable'), self.tr('Improved'), self.tr('No Data')]
        y = [get_xtab_area(trans_prod_xtab, -1, None),
             get_xtab_area(trans_prod_xtab, 0, None),
             get_xtab_area(trans_prod_xtab, 1, None),
             get_xtab_area(trans_prod_xtab, -32768, None)]
        log('SummaryTable total area: {}'.format(sum(y)))
        log('SummaryTable areas (deg, stable, imp, no data): {}'.format(y))

        make_summary_table(soc_bl_totals, soc_tg_totals, trans_prod_xtab,
                           self.output_file_table.text())

        style_sdg_ld(prod_f, QtGui.QApplication.translate('LDMPPlugin', 'SDG 15.3.1 productivity sub-indicator'))
        style_sdg_ld(self.output_file_layer.text(), QtGui.QApplication.translate('LDMPPlugin', 'Degradation (SDG 15.3.1 indicator)'))

        self.plot_degradation(x, y)

    def plot_degradation(self, x, y):
        dlg_plot = DlgPlotBars()
        labels = {'title': self.plot_title.text(),
                  'bottom': self.tr('Land status'),
                  'left': [self.tr('Area'), self.tr('km<sup>2</sup>')]}
        dlg_plot.plot_data(x, y, labels)
        dlg_plot.show()
        dlg_plot.exec_()


def get_lc_area(table, code):
    ind = np.where(table[0] == code)[0]
    if ind.size == 0:
        return 0
    else:
        return float(table[1][ind])


def get_prod_table(table, change_type, classes=range(1, 7 + 1)):
    out = np.zeros((len(classes), len(classes)))
    for bl_class in range(len(classes)):
        for tg_class in range(len(classes)):
            transition = int('{}{}'.format(classes[bl_class], classes[tg_class]))
            if change_type == 'improved':
                out[bl_class, tg_class] = get_xtab_area(table, 1, transition)
            elif change_type == 'stable':
                out[bl_class, tg_class] = get_xtab_area(table, 0, transition)
            elif change_type == 'degraded':
                out[bl_class, tg_class] = get_xtab_area(table, -1, transition)
            if change_type == 'no data':
                out[bl_class, tg_class] = get_xtab_area(table, -32768, transition)
    return out


# Note classes for this function go from 1-6 to exclude water from the SOC 
# totals
def write_soc_stock_change_table(sheet, first_row, first_col, soc_bl_totals, 
                                 soc_tg_totals, classes=range(1, 6 + 1)):
    for row in range(len(classes)):
        for col in range(len(classes)):
            cell = sheet.cell(row=row + first_row, column=col + first_col)
            transition = int('{}{}'.format(classes[row], classes[col]))
            bl_soc = get_soc_total(soc_bl_totals, transition)
            tg_soc = get_soc_total(soc_tg_totals, transition)
            cell.value = tg_soc - bl_soc
    

# Note classes for this function go from 1-6 to exclude water from the SOC 
# totals
def get_soc_bl_tg(trans_prod_xtab, soc_bl_totals, soc_tg_totals, classes=range(1, 6 + 1)):
    out = np.zeros((len(classes), 2))
    for row in range(len(classes)):
        bl_area = 0
        bl_soc = 0
        tg_area = 0
        tg_soc = 0
        # Need to sum up the total soc across the pixels and then divide by 
        # total area
        for n in range(len(classes)):
            bl_trans = int('{}{}'.format(classes[row], classes[n]))
            bl_area += get_xtab_area(trans_prod_xtab, None, bl_trans)
            bl_soc += get_soc_total(soc_bl_totals, bl_trans)

            tg_trans = int('{}{}'.format(classes[n], classes[row]))
            tg_area += get_xtab_area(trans_prod_xtab, None, tg_trans)
            tg_soc += get_soc_total(soc_tg_totals, tg_trans)
        # Note areas are in sq km. Need to convert to ha
        if bl_soc != 0 and bl_area != 0:
            out[row][0] = bl_soc / (bl_area * 100)
        else:
            out[row][0]
        if tg_soc != 0 and tg_area != 0:
            out[row][1] = tg_soc / (tg_area * 100)
        else:
            out[row][1] = 0
    return out


def get_lc_table(table, classes=range(1, 7 + 1)):
    out = np.zeros((len(classes), len(classes)))
    for bl_class in range(len(classes)):
        for tg_class in range(len(classes)):
            transition = int('{}{}'.format(classes[bl_class], classes[tg_class]))
            out[bl_class, tg_class] = get_xtab_area(table, None, transition)
    return out


def get_soc_total(soc_table, transition):
    ind = np.where(soc_table[0] == transition)[0]
    if ind.size == 0:
        return 0
    else:
        return float(soc_table[1][ind])


def write_table_to_sheet(sheet, d, first_row, first_col):
    for row in range(d.shape[0]):
        for col in range(d.shape[1]):
            cell = sheet.cell(row=row + first_row, column=col + first_col)
            cell.value = d[row, col]


def make_summary_table(soc_bl_totals, soc_tg_totals, trans_prod_xtab, out_file):
    def tr(s):
        return QtGui.QApplication.translate("LDMP", s)

    wb = openpyxl.load_workbook(os.path.join(os.path.dirname(os.path.realpath(__file__)), 'data', 'SummaryTable.xlsx'))

    ##########################################################################
    # Productivity tables
    ws_prod = wb.get_sheet_by_name('Productivity')

    write_table_to_sheet(ws_prod, get_prod_table(trans_prod_xtab, 'improved'), 14, 3)
    write_table_to_sheet(ws_prod, get_prod_table(trans_prod_xtab, 'stable'), 26, 3)
    write_table_to_sheet(ws_prod, get_prod_table(trans_prod_xtab, 'degraded'), 38, 3)
    write_table_to_sheet(ws_prod, get_prod_table(trans_prod_xtab, 'no data'), 50, 3)

    ##########################################################################
    # Soil organic carbon tables
    ws_soc = wb.get_sheet_by_name('Soil organic carbon')

    write_table_to_sheet(ws_soc, get_soc_bl_tg(trans_prod_xtab, soc_bl_totals, soc_tg_totals), 8, 3)

    # Write table of baseline areas
    lc_trans_table_no_water = get_lc_table(trans_prod_xtab, classes=np.arange(1, 6 + 1))
    write_table_to_sheet(ws_soc, np.reshape(np.sum(lc_trans_table_no_water, 1), (-1, 1)), 8, 5)
    # Write table of target areas
    write_table_to_sheet(ws_soc, np.reshape(np.sum(lc_trans_table_no_water, 0), (-1, 1)), 8, 6)
    
    # write_soc_stock_change_table has its own writing function as it needs to write a 
    # mix of numbers and strings
    write_soc_stock_change_table(ws_soc, 19, 3, soc_bl_totals, soc_tg_totals)

    ##########################################################################
    # Land cover tables
    ws_lc = wb.get_sheet_by_name('Land cover')
    write_table_to_sheet(ws_lc, get_lc_table(trans_prod_xtab), 18, 3)

    ws_prod_logo = Image(os.path.join(os.path.dirname(__file__), 'data', 'trends_earth_logo_bl_300width.png'))
    ws_prod.add_image(ws_prod_logo, 'H1')
    ws_soc_logo = Image(os.path.join(os.path.dirname(__file__), 'data', 'trends_earth_logo_bl_300width.png'))
    ws_soc.add_image(ws_soc_logo, 'H1')
    ws_lc_logo = Image(os.path.join(os.path.dirname(__file__), 'data', 'trends_earth_logo_bl_300width.png'))
    ws_lc.add_image(ws_lc_logo, 'H1')

    try:
        wb.save(out_file)
        log('Indicator table saved to {}'.format(out_file))
        # QtGui.QMessageBox.information(None, QtGui.QApplication.translate("LDMP", "Success"),
        #         QtGui.QApplication.translate("LDMP", 'Indicator table saved to <a href="file://{}">{}</a>'.format(out_file, out_file)))
        QtGui.QMessageBox.information(None, QtGui.QApplication.translate("LDMP", "Success"),
                                      QtGui.QApplication.translate("LDMP", 'Indicator table saved to {}'.format(out_file)))

    except IOError:
        log('Error saving {}'.format(out_file))
        QtGui.QMessageBox.critical(None, QtGui.QApplication.translate("LDMP", "Error"),
                                   QtGui.QApplication.translate("LDMP", "Error saving output table - check that {} is accessible and not already open.".format(out_file)), None)


class DlgCreateMap(DlgCalculateBase, Ui_DlgCreateMap):
    '''Class to be shared across SDG and SummaryTable reporting dialogs'''

    def __init__(self, parent=None):
        super(DlgCreateMap, self).__init__(parent)
        self.setupUi(self)

    def firstShow(self):
        #TODO: Remove the combo page for now...
        self.combo_layers.hide()
        self.layer_combo_label.hide()
        self.TabBox.removeTab(1)

        super(DlgCreateMap, self).firstShow()

    def showEvent(self, event):
        super(DlgCreateMap, self).showEvent(event)

        QtGui.QMessageBox.warning(None, QtGui.QApplication.translate("LDMP", "Warning"),
                                  QtGui.QApplication.translate("LDMP", "The create map tool is still experimental - the functionality of this tool is likely to change in the future."), None)

        self.populate_layers()

    def populate_layers(self):
        self.combo_layers.clear()
        self.layers_list = get_ld_layers()
        self.combo_layers.addItems([l[0].name() for l in self.layers_list])

    def btn_calculate(self):
        # Note that the super class has several tests in it - if they fail it
        # returns False, which would mean this function should stop execution
        # as well.

        #TODO Will need to reenable this if the area combo selector is used in the future
        # ret = super(DlgCreateMap, self).btn_calculate()
        # if not ret:
        #     return

        self.close()

        if self.portrait_layout.isChecked():
            orientation = 'portrait'
        else:
            orientation = 'landscape'

        template = os.path.join(os.path.dirname(__file__), 'data',
                                'map_template_{}.qpt'.format(orientation))

        f = file(template, 'rt')
        new_composer_content = f.read()
        f.close()
        document = QtXml.QDomDocument()
        document.setContent(new_composer_content)

        if self.title.text():
            title = self.title.text()
        else:
            title = 'trends.earth map'
        comp_window = iface.createNewComposer(title)
        composition = comp_window.composition()
        composition.loadFromTemplate(document)

        canvas = iface.mapCanvas()
        map_item = composition.getComposerItemById('te_map')
        map_item.setMapCanvas(canvas)
        map_item.zoomToExtent(canvas.extent())

        # Add area of interest
        # layerset = []
        # aoi_layer = QgsVectorLayer("Polygon?crs=epsg:4326", "Area of interest", "memory")
        # mask_pr = aoi_layer.dataProvider()
        # fet = QgsFeature()
        # fet.setGeometry(self.aoi)
        # mask_pr.addFeatures([fet])
        # QgsMapLayerRegistry.instance().addMapLayer(aoi_layer)
        # layerset.append(aoi_layer.id())
        # map_item.setLayerSet(layerset)
        # map_item.setKeepLayerSet(True)

        map_item.renderModeUpdateCachedImage()

        datasets = composition.getComposerItemById('te_datasets')
        datasets.setText('Created using <a href="http://trends.earth">trends.earth</a>. Projection: decimal degrees, WGS84. Datasets derived from {{COMING SOON}}.')
        datasets.setHtmlState(True)
        author = composition.getComposerItemById('te_authors')
        author.setText(self.authors.text())
        logo = composition.getComposerItemById('te_logo')
        logo_path = os.path.join(os.path.dirname(__file__), 'data', 'trends_earth_logo_bl_600width.png')
        logo.setPicturePath(logo_path)
        legend = composition.getComposerItemById('te_legend')
        legend.setAutoUpdateModel(True)
